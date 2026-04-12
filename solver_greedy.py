#!/usr/bin/env python3
# LineToolingOptimization Solver
# Reads inputs from Blad1, optimizes, writes results back.
# Usage: python solver.py [workbook.xlsx]
# Requirements: pip install openpyxl
import sys, math, argparse
from pathlib import Path
import openpyxl
SHEET_NAME, SUMMARY_SHEET = 'Blad1', 'Summary'
NUM_PRODUCTS, NUM_LINES, LOOK_AHEAD_YEARS = 10, 15, 3
DEMAND_START, DEMAND_END = 3, 19
CT_START, OEE_START = 31, 49
MECH_START, OPT_START = 70, 84
OUTPUT_START = 98
def load_inputs(wb):
    ws = wb[SHEET_NAME]
    setup = {
        'hours': float(ws['B22'].value or 0),
        'shifts': float(ws['B23'].value or 0),
        'days': float(ws['B24'].value or 0),
        'weeks': float(ws['B25'].value or 0),
    }
    avail = setup['hours']*setup['shifts']*setup['days']*setup['weeks']*3600
    years, demand = [], {}
    for r in range(DEMAND_START, DEMAND_END + 1):
        yr = ws.cell(row=r, column=2).value
        if yr is None: continue
        yr = int(yr)
        years.append(yr)
        demand[yr] = [int(ws.cell(row=r, column=c).value or 0) for c in range(3, 13)]
    ct = [[ws.cell(row=CT_START+l, column=3+p).value or 12 for p in range(10)] for l in range(15)]
    oee = [[ws.cell(row=OEE_START+l, column=3+p).value or 0.85 for p in range(10)] for l in range(15)]
    mt = [[ws.cell(row=MECH_START+i, column=3+j).value or 0 for j in range(10)] for i in range(10)]
    ot = [[ws.cell(row=OPT_START+i, column=3+j).value or 0 for j in range(10)] for i in range(10)]
    return {'years':years,'demand':demand,'avail':avail,'ct':ct,'oee':oee,'mt':mt,'ot':ot}
def eff_sec(ct_val, oee_val):
    return ct_val / oee_val if oee_val > 0 else float('inf')
def solve(inp):
    years, demand, avail = inp['years'], inp['demand'], inp['avail']
    ct, oee_m = inp['ct'], inp['oee']
    alloc = {}  # (year, line, product) -> units
    tooled = [set() for _ in range(NUM_LINES)]
    intro = {}
    def es(l, p): return eff_sec(ct[l][p], oee_m[l][p])
    def mu(l, p): return int(avail / es(l, p))
    for yr in years:
        d = demand.get(yr, [0]*10)
        active = [p for p in range(10) if d[p] > 0]
        if not active: continue
        rem = list(d)
        ltu = [0.0] * NUM_LINES  # line time used
        td = sum(d[p] for p in active)
        tt = sum(d[p] * es(0, p) for p in active)
        # Line 1: Validation
        if tt <= avail:
            for p in active:
                alloc[(yr,0,p)] = rem[p]; ltu[0] += rem[p]*es(0,p); rem[p] = 0
            tooled[0].update(active)
            if 0 not in intro: intro[0] = yr
            continue
        dom = max(active, key=lambda p: d[p])
        small = [p for p in active if p != dom]
        small_t = sum(d[p] * es(0, p) for p in small)
        if small_t <= avail:
            for p in small:
                alloc[(yr,0,p)] = rem[p]; ltu[0] += rem[p]*es(0,p); rem[p] = 0
            tl = avail - ltu[0]
            u = min(rem[dom], int(tl / es(0, dom)))
            if u > 0:
                alloc[(yr,0,dom)] = u; ltu[0] += u*es(0,dom); rem[dom] -= u
            tooled[0].update(active)
        else:
            for p in active:
                prop = d[p] / td
                u = min(rem[p], int((prop * avail) / es(0, p)))
                if u > 0:
                    alloc[(yr,0,p)] = u; ltu[0] += u*es(0,p); rem[p] -= u
            tooled[0].update(active)
        if 0 not in intro: intro[0] = yr
        # Lines 2+: Overflow
        # Phase A: Reuse tooled lines
        for p in sorted(range(10), key=lambda p: -rem[p]):
            if rem[p] <= 0: continue
            for line in range(1, NUM_LINES):
                if rem[p] <= 0: break
                if p not in tooled[line]: continue
                tl = avail - ltu[line]
                if tl < es(line, p): continue
                a = min(rem[p], int(tl / es(line, p)))
                alloc[(yr,line,p)] = alloc.get((yr,line,p),0) + a
                ltu[line] += a * es(line, p); rem[p] -= a
        # Phase B: Pack or open new lines
        items = [(p, rem[p]) for p in range(10) if rem[p] > 0]
        items.sort(key=lambda x: -x[1])
        for p, vol in items:
            while vol > 0:
                bl, bs = -1, -999999
                for line in range(1, NUM_LINES):
                    tl = avail - ltu[line]
                    if tl < es(line, p): continue
                    s = 0
                    if p in tooled[line]: s += 1000000
                    if ltu[line] > 0: s += 500000 + int(ltu[line]/1000)
                    s -= line
                    if s > bs: bl, bs = line, s
                if bl == -1:
                    print(f'  Warning: {yr} P{p+1} cannot place {vol:,}')
                    break
                mx = int((avail - ltu[bl]) / es(bl, p))
                a = min(vol, mx)
                alloc[(yr,bl,p)] = alloc.get((yr,bl,p),0) + a
                ltu[bl] += a * es(bl, p); vol -= a
                tooled[bl].add(p)
                if bl not in intro: intro[bl] = yr
    # Look-ahead: pre-add tooling during intro year
    for line in sorted(intro.keys()):
        iy = intro[line]
        for fy in range(iy+1, min(iy+LOOK_AHEAD_YEARS+1, max(years)+1)):
            for p in range(10):
                if alloc.get((fy, line, p), 0) > 0:
                    tooled[line].add(p)
    return alloc, tooled, intro
def write_output(ws, alloc, years):
    for r in range(OUTPUT_START, OUTPUT_START + len(years)):
        for c in range(3, 3 + NUM_LINES * NUM_PRODUCTS):
            ws.cell(row=r, column=c).value = None
    n = 0
    for yr in years:
        row = OUTPUT_START + (yr - years[0])
        for line in range(NUM_LINES):
            for prod in range(NUM_PRODUCTS):
                v = alloc.get((yr, line, prod), 0)
                if v > 0:
                    ws.cell(row=row, column=3+line*10+prod).value = v
                    n += 1
    return n
def verify(alloc, inp):
    ok = True
    for yr in inp['years']:
        d = inp['demand'].get(yr, [0]*10)
        if sum(d) == 0: continue
        ta = [0]*10
        for l in range(NUM_LINES):
            for p in range(10): ta[p] += alloc.get((yr,l,p), 0)
        for p in range(10):
            if abs(d[p] - ta[p]) > 1:
                print(f'  X {yr} P{p+1}: demand={d[p]:,} alloc={ta[p]:,}')
                ok = False
        for l in range(NUM_LINES):
            t = sum(alloc.get((yr,l,p),0)*eff_sec(inp['ct'][l][p],inp['oee'][l][p]) for p in range(10))
            if t > inp['avail'] * 1.001:
                print(f'  X {yr} Line {l+1}: time exceeded')
                ok = False
    return ok
def main():
    parser = argparse.ArgumentParser(description='Production line optimizer')
    parser.add_argument('workbook', nargs='?', default='LineToolingOptimization.xlsx')
    args = parser.parse_args()
    path = Path(args.workbook)
    if not path.exists(): print(f'Error: {path} not found'); sys.exit(1)
    print(f'Loading {path} ...')
    wb = openpyxl.load_workbook(str(path))
    inp = load_inputs(wb)
    print(f'  Avail seconds: {inp["avail"]:,.0f}')
    print('Running solver ...')
    alloc, tooled, intro = solve(inp)
    print('Verifying ...')
    ok = verify(alloc, inp)
    print(f'  {"All OK" if ok else "Issues found"}')
    tm = sum(len(s) for s in tooled)
    print(f'Tooling: {tm} mech + {tm} opt = {tm*2} sets')
    for l in sorted(intro.keys()):
        ps = sorted([f"P{p+1}" for p in tooled[l]])
        print(f'  Line {l+1} ({intro[l]}): {ps}')
    ws = wb[SHEET_NAME]
    n = write_output(ws, alloc, inp['years'])
    print(f'Wrote {n} cells to output grid')
    print(f'Saving {path} ...')
    wb.save(str(path))
    print('Done!')
if __name__ == '__main__':
    main()

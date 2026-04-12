#!/usr/bin/env python3
"""
create_simple_xlsx.py
Generates Simple.xlsx — a lightweight workbook for solver_simple.py.

Sheets: Demand · Parameters · Allocation

No tooling matrices. Two cost levers only: annual running cost and changeover
penalty. Everything else (OEE, cycle times, demand) is the same as Solver.xlsx.

If Solver.xlsx (or any Solver-format workbook) is supplied, demand, schedule,
cycle times, and OEE are copied from it automatically.

Usage:
    python create_simple_xlsx.py               # use built-in defaults
    python create_simple_xlsx.py Solver.xlsx   # copy inputs from Solver.xlsx
"""
import sys
import argparse
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════════
# LAYOUT CONSTANTS  — must match the corresponding values in solver_simple.py
# ═══════════════════════════════════════════════════════════════════════════════

# 'Demand' sheet  (identical layout to Solver.xlsx)
DEM_HDR_ROW  = 4
DEM_DATA_ROW = 5
# col 1=Year  col 2=P1 … col 11=P10  col 12=Total

# 'Parameters' sheet
PAR_HOURS_ROW    = 4
PAR_SHIFTS_ROW   = 5
PAR_DAYS_ROW     = 6
PAR_WEEKS_ROW    = 7
PAR_VAL_COL      = 2
PAR_CT_HDR_ROW   = 13
PAR_CT_DATA_ROW  = 14   # 15 rows (Line 1 … Line 15)
PAR_OEE_HDR_ROW  = 33   # col-header row for OEE matrix (row 31=sect, 32=note, 33=hdr)
PAR_OEE_DATA_ROW = 34   # first OEE data row (Line 1)
# Cost section (simplified — only 2 rows)
PAR_COST_HDR_ROW  = 50
PAR_COST_RUN_ROW  = 51  # annual running cost per open line
PAR_COST_CHGOVER_ROW = 52  # penalty per mixed-product line-year

# 'Allocation' sheet  (written by solver_simple.py)
ALLOC_HDR_ROW  = 5
ALLOC_DATA_ROW = 6

YEARS    = list(range(2025, 2042))
PRODUCTS = [f'P{i+1}' for i in range(10)]
LINES    = [f'L{i+1}' for i in range(15)]

# ═══════════════════════════════════════════════════════════════════════════════
# COLOUR PALETTE & STYLE HELPERS  (same palette as create_solver_xlsx.py)
# ═══════════════════════════════════════════════════════════════════════════════
C_TITLE   = '1F3864'
C_SECT    = '2E75B6'
C_COLHDR  = 'BDD7EE'
C_INPUT   = 'FFF9C4'
C_RONLY   = 'F2F2F2'
C_OUTPUT  = 'E2EFDA'
C_AMBER   = 'FFE699'   # mixed-product line indicator
C_WHITE   = 'FFFFFF'
C_BORDER  = 'B8CCE4'


def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)


def _font(bold=False, italic=False, size=11, color='1F1F1F'):
    return Font(name='Calibri', bold=bold, italic=italic, size=size, color=color)


_THIN = Side(style='thin',   color=C_BORDER)
_MED  = Side(style='medium', color='4472C4')


def _border_thin():
    return Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _al(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _w(ws, row, col, value=None, *,
       bold=False, italic=False, sz=11, fc='1F1F1F',
       bg=None, h='left', v='center', wrap=False, bdr=True, fmt=None):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    c.font      = _font(bold=bold, italic=italic, size=sz, color=fc)
    c.alignment = _al(h=h, v=v, wrap=wrap)
    if bg:
        c.fill = _fill(bg)
    if bdr:
        c.border = _border_thin()
    if fmt:
        c.number_format = fmt
    return c


def _title(ws, row, text, ncols=12, sz=14):
    _w(ws, row, 1, text, bold=True, sz=sz, fc=C_WHITE, bg=C_TITLE,
       h='center', bdr=False)
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row,   end_column=ncols)
    ws.row_dimensions[row].height = 28


def _sect(ws, row, text, ncols=12):
    _w(ws, row, 1, text, bold=True, fc=C_WHITE, bg=C_SECT, bdr=False)
    for col in range(2, ncols + 1):
        ws.cell(row=row, column=col).fill = _fill(C_SECT)
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row,   end_column=ncols)
    ws.row_dimensions[row].height = 20


def _col_headers(ws, row, labels, start_col=1):
    for i, lbl in enumerate(labels):
        _w(ws, row, start_col + i, lbl, bold=True,
           bg=C_COLHDR, h='center', bdr=True)
    ws.row_dimensions[row].height = 18


def _note(ws, row, text, ncols=12):
    c = ws.cell(row=row, column=1)
    c.value = text
    c.font  = _font(italic=True, size=9, color='808080')
    c.alignment = _al(wrap=True)
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row,   end_column=ncols)
    ws.row_dimensions[row].height = 15


# ═══════════════════════════════════════════════════════════════════════════════
# SOURCE-DATA READING  (copy from Solver.xlsx if available)
# ═══════════════════════════════════════════════════════════════════════════════

def _read_solver_xlsx(wb):
    """Read demand, schedule, cycle times, OEE from a Solver.xlsx workbook."""
    ws_d = wb['Demand']
    ws_p = wb['Parameters']

    setup = {
        'hours':  float(ws_p.cell(row=PAR_HOURS_ROW,  column=PAR_VAL_COL).value or 7.2),
        'shifts': float(ws_p.cell(row=PAR_SHIFTS_ROW, column=PAR_VAL_COL).value or 3),
        'days':   float(ws_p.cell(row=PAR_DAYS_ROW,   column=PAR_VAL_COL).value or 6),
        'weeks':  float(ws_p.cell(row=PAR_WEEKS_ROW,  column=PAR_VAL_COL).value or 48),
    }

    demand = {}
    for r in range(DEM_DATA_ROW, DEM_DATA_ROW + 50):
        yr = ws_d.cell(row=r, column=1).value
        if yr is None:
            break
        demand[int(yr)] = [int(ws_d.cell(row=r, column=2 + p).value or 0)
                           for p in range(10)]

    ct  = [[ws_p.cell(row=PAR_CT_DATA_ROW  + l, column=2 + p).value or 12
            for p in range(10)] for l in range(15)]
    oee = [[ws_p.cell(row=PAR_OEE_DATA_ROW + l, column=2 + p).value or 0.85
            for p in range(10)] for l in range(15)]

    return {'setup': setup, 'demand': demand, 'ct': ct, 'oee': oee}


def _defaults():
    demand = {yr: [0] * 10 for yr in YEARS}
    for yr, d in [
        (2029, [1000,    1000,    1000,    1000,    0,0,0,0,0,0]),
        (2030, [35498,   141992,  35498,   35498,   0,0,0,0,0,0]),
        (2031, [372560,  1490238, 372560,  372560,  0,0,0,0,0,0]),
        (2032, [529763,  2119050, 529763,  529763,  0,0,0,0,0,0]),
        (2033, [693177,  2772708, 693177,  693177,  0,0,0,0,0,0]),
        (2034, [1001694, 4006776, 1001694, 1001694, 0,0,0,0,0,0]),
        (2035, [1338697, 5354788, 1338697, 1338697, 0,0,0,0,0,0]),
        (2036, [1286832, 5417328, 1286832, 1286832, 0,0,0,0,0,0]),
        (2037, [610237,  2440948, 610237,  610237,  0,0,0,0,0,0]),
        (2038, [173752,  695008,  173752,  173752,  0,0,0,0,0,0]),
    ]:
        demand[yr] = d
    return {
        'setup':  {'hours': 7.2, 'shifts': 3, 'days': 6, 'weeks': 48},
        'demand': demand,
        'ct':  [[12.0] * 10 for _ in range(15)],
        'oee': [[0.85] * 10 for _ in range(15)],
    }


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET: Demand
# ═══════════════════════════════════════════════════════════════════════════════

def _build_demand(wb, data):
    ws = wb.create_sheet('Demand')
    ws.sheet_properties.tabColor = '70AD47'
    ws.sheet_view.showGridLines  = False
    ws.freeze_panes = 'B5'

    ws.column_dimensions['A'].width = 8
    for col in range(2, 12):
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.column_dimensions[get_column_letter(12)].width = 15

    _title(ws, 1, 'VOLUME DEMAND', ncols=12)
    _note(ws, 2, 'Annual unit demand per product  ·  Edit yellow cells  ·  '
          'P5-P10 reserved for future products (leave 0 until needed)', ncols=12)
    ws.row_dimensions[3].height = 6
    _col_headers(ws, DEM_HDR_ROW, ['Year'] + PRODUCTS + ['Total'])

    demand = data['demand']
    for idx, yr in enumerate(YEARS):
        row = DEM_DATA_ROW + idx
        d   = demand.get(yr, [0] * 10)
        _w(ws, row, 1, yr, bold=True, bg=C_RONLY, h='center')
        for p in range(10):
            val = d[p] if d[p] != 0 else None
            _w(ws, row, 2 + p, val, bg=C_INPUT, h='right', fmt='#,##0')
        first = get_column_letter(2)
        last  = get_column_letter(11)
        _w(ws, row, 12, f'=SUM({first}{row}:{last}{row})',
           bg=C_RONLY, h='right', fmt='#,##0')
        ws.row_dimensions[row].height = 16

    _note(ws, DEM_DATA_ROW + len(YEARS) + 1,
          'Demand of 0 means the product is inactive that year — the solver ignores it.',
          ncols=12)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET: Parameters
# ═══════════════════════════════════════════════════════════════════════════════

def _build_parameters(wb, data):
    ws = wb.create_sheet('Parameters')
    ws.sheet_properties.tabColor = 'ED7D31'
    ws.sheet_view.showGridLines  = False
    ws.freeze_panes = 'C14'

    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 12
    for col in range(3, 13):
        ws.column_dimensions[get_column_letter(col)].width = 10

    _title(ws, 1, 'PRODUCTION PARAMETERS', ncols=12)
    ws.row_dimensions[2].height = 6

    # ── Setup ──────────────────────────────────────────────────────────────────
    _sect(ws, 3, '  PRODUCTION SETUP', ncols=3)
    s = data['setup']
    setup_rows = [
        (PAR_HOURS_ROW,  'Hours per shift',         s['hours'],  '0.0',   'hrs',    True),
        (PAR_SHIFTS_ROW, 'Shifts per day',           s['shifts'], '0',     'shifts', True),
        (PAR_DAYS_ROW,   'Working days per week',    s['days'],   '0',     'days',   True),
        (PAR_WEEKS_ROW,  'Working weeks per year',   s['weeks'],  '0',     'weeks',  True),
        (8,  'Available seconds / year', '=B4*B5*B6*B7*3600', '#,##0', 'sec/yr', False),
        (9,  'Available hours / year',   '=B8/3600',          '#,##0', 'hrs/yr', False),
    ]
    for row, label, val, fmt, unit, editable in setup_rows:
        ws.cell(row=row, column=1).value     = label
        ws.cell(row=row, column=1).font      = _font(size=10)
        ws.cell(row=row, column=1).fill      = _fill(C_RONLY)
        ws.cell(row=row, column=1).border    = _border_thin()
        ws.cell(row=row, column=1).alignment = _al()
        _w(ws, row, 2, val, bold=editable,
           bg=C_INPUT if editable else C_RONLY, h='right', fmt=fmt)
        ws.cell(row=row, column=3).value = unit
        ws.cell(row=row, column=3).font  = _font(size=9, color='808080')
        ws.row_dimensions[row].height    = 18

    ws.row_dimensions[10].height = 8

    # ── Cycle Time ─────────────────────────────────────────────────────────────
    _sect(ws, 11, '  CYCLE TIME  (seconds per unit)', ncols=12)
    _note(ws, 12, 'Default 12 s/unit — adjust per product / line as needed.',
          ncols=12)
    _col_headers(ws, PAR_CT_HDR_ROW, ['Line'] + PRODUCTS)

    ct = data['ct']
    for l in range(15):
        row = PAR_CT_DATA_ROW + l
        _w(ws, row, 1, LINES[l], bold=True, bg=C_RONLY, h='center')
        for p in range(10):
            _w(ws, row, 2 + p, ct[l][p], bg=C_INPUT, h='right', fmt='0.0')
        ws.row_dimensions[row].height = 16

    ws.row_dimensions[PAR_CT_DATA_ROW + 15].height = 8

    # ── OEE ────────────────────────────────────────────────────────────────────
    _sect(ws, PAR_OEE_HDR_ROW - 2, '  BASE OEE  (single-product line)', ncols=12)
    _note(ws, PAR_OEE_HDR_ROW - 1,
          'Overall Equipment Effectiveness per line per product.  '
          'Multi-product lines incur a 3% penalty (hardcoded in solver_simple.py).',
          ncols=12)
    _col_headers(ws, PAR_OEE_HDR_ROW, ['Line'] + PRODUCTS)

    oee = data['oee']
    for l in range(15):
        row = PAR_OEE_DATA_ROW + l
        _w(ws, row, 1, LINES[l], bold=True, bg=C_RONLY, h='center')
        for p in range(10):
            _w(ws, row, 2 + p, oee[l][p], bg=C_INPUT, h='right', fmt='0.00')
        ws.row_dimensions[row].height = 16

    ws.row_dimensions[PAR_OEE_DATA_ROW + 15].height = 8

    # ── Cost parameters ────────────────────────────────────────────────────────
    _sect(ws, PAR_COST_HDR_ROW, '  COST PARAMETERS', ncols=3)
    cost_rows = [
        (PAR_COST_RUN_ROW,    'Annual running cost per open line',
         500_000, '$#,##0', 'USD / line / year'),
        (PAR_COST_CHGOVER_ROW, 'Changeover penalty (mixed-product line-year)',
         100_000, '$#,##0', 'USD / mixed-line / year'),
    ]
    for row, label, val, fmt, unit in cost_rows:
        ws.cell(row=row, column=1).value     = label
        ws.cell(row=row, column=1).font      = _font(size=10)
        ws.cell(row=row, column=1).fill      = _fill(C_RONLY)
        ws.cell(row=row, column=1).border    = _border_thin()
        ws.cell(row=row, column=1).alignment = _al()
        _w(ws, row, 2, val, bold=True, bg=C_INPUT, h='right', fmt=fmt)
        ws.cell(row=row, column=3).value = unit
        ws.cell(row=row, column=3).font  = _font(size=9, color='808080')
        ws.row_dimensions[row].height    = 18

    _note(ws, PAR_COST_CHGOVER_ROW + 1,
          'Running cost discourages opening lines early. '
          'Changeover penalty encourages dedicating lines to single products.',
          ncols=3)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET: Allocation  (placeholder — overwritten by solver_simple.py)
# ═══════════════════════════════════════════════════════════════════════════════

def _build_allocation(wb):
    ws = wb.create_sheet('Allocation')
    ws.sheet_properties.tabColor = 'A9D18E'
    ws.sheet_view.showGridLines  = False

    _title(ws, 1, 'ALLOCATION  —  solver output', ncols=20)
    _note(ws, 2,
          'This sheet is overwritten each time solver_simple.py runs.  '
          'Run: python solver_simple.py Simple.xlsx',
          ncols=20)
    ws.row_dimensions[3].height = 6
    ws.row_dimensions[4].height = 6

    _note(ws, 5,
          'No results yet — run solver_simple.py to populate this sheet.',
          ncols=20)


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description='Generate Simple.xlsx for solver_simple.py')
    parser.add_argument('source', nargs='?',
                        help='Solver.xlsx to copy inputs from (optional)')
    args = parser.parse_args()

    # Load source data
    if args.source:
        src = Path(args.source)
        if not src.exists():
            print(f'ERROR: source file not found: {src}')
            sys.exit(1)
        print(f'Reading inputs from {src} ...')
        src_wb = openpyxl.load_workbook(str(src), data_only=True)
        data = _read_solver_xlsx(src_wb)
        print('  Demand, schedule, cycle times, OEE copied.')
    else:
        print('No source file — using built-in defaults.')
        data = _defaults()

    # Build workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default Sheet

    _build_demand(wb, data)
    _build_parameters(wb, data)
    _build_allocation(wb)

    out = Path('Simple.xlsx')
    wb.save(str(out))
    print(f'\nSaved {out}')
    print('Next step: python solver_simple.py Simple.xlsx')


if __name__ == '__main__':
    main()

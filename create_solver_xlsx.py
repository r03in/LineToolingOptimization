#!/usr/bin/env python3
"""
create_solver_xlsx.py
Generates Solver.xlsx — a clean, formatted workbook for LineToolingOptimization.

Sheets: Instructions · Demand · Parameters · Tooling · Allocation · Report

If Book.xlsx (legacy) is supplied, input data is copied from it;
otherwise built-in defaults are used.

Usage:
    python create_solver_xlsx.py
    python create_solver_xlsx.py Book.xlsx

Layout row/column constants at the top are kept in sync with solver_ilp.py.
Requirements: pip install openpyxl
"""
import sys
import argparse
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════════
# LAYOUT CONSTANTS  — must match the corresponding values in solver_ilp.py
# ═══════════════════════════════════════════════════════════════════════════════

# 'Demand' sheet
DEM_HDR_ROW  = 4   # column-header row   (Year | P1 … P10 | Total)
DEM_DATA_ROW = 5   # first data row       (year 2025)
# col 1=Year  col 2=P1  …  col 11=P10  col 12=Total

# 'Parameters' sheet
PAR_HOURS_ROW    = 4    # hours / shift
PAR_SHIFTS_ROW   = 5    # shifts / day
PAR_DAYS_ROW     = 6    # days / week
PAR_WEEKS_ROW    = 7    # weeks / year
PAR_VAL_COL      = 2    # values in col B
PAR_CT_HDR_ROW   = 13   # col-header row for cycle-time matrix
PAR_CT_DATA_ROW  = 14   # first CT data row  (Line 1)
PAR_OEE_HDR_ROW  = 32   # col-header row for OEE matrix
PAR_OEE_DATA_ROW = 33   # first OEE data row (Line 1)
# col 1=Line-label  col 2=P1  …  col 11=P10

# 'Tooling' sheet
TOO_MECH_HDR_ROW  = 6   # col-header row, mech matrix
TOO_MECH_DATA_ROW = 7   # first mech data row (P1)
TOO_OPT_HDR_ROW   = 20  # col-header row, optical matrix
TOO_OPT_DATA_ROW  = 21  # first optical data row (P1)
# col 1=Product-label  col 2=P1  …  col 11=P10

# 'Allocation' sheet  (written by solver)
ALLOC_HDR_ROW  = 5
ALLOC_DATA_ROW = 6

YEARS    = list(range(2025, 2042))   # 17 years
PRODUCTS = [f'P{i+1}' for i in range(10)]
LINES    = [f'L{i+1}' for i in range(15)]

# ═══════════════════════════════════════════════════════════════════════════════
# COLOUR PALETTE & STYLE HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
C_TITLE   = '1F3864'   # dark navy  – sheet titles
C_SECT    = '2E75B6'   # medium blue – section banners
C_COLHDR  = 'BDD7EE'   # light blue  – column headers
C_INPUT   = 'FFF9C4'   # pale yellow – editable input
C_RONLY   = 'F2F2F2'   # light grey  – read-only / computed
C_OUTPUT  = 'E2EFDA'   # light green – solver output
C_WHITE   = 'FFFFFF'
C_BORDER  = 'B8CCE4'


def _fill(hex_color: str) -> PatternFill:
    return PatternFill('solid', fgColor=hex_color)


def _font(bold=False, italic=False, size=11, color='1F1F1F') -> Font:
    return Font(name='Calibri', bold=bold, italic=italic,
                size=size, color=color)


_THIN = Side(style='thin',   color=C_BORDER)
_MED  = Side(style='medium', color='4472C4')


def _border_thin() -> Border:
    return Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _al(h='left', v='center', wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _w(ws, row: int, col: int, value=None, *,
       bold=False, italic=False, sz=11, fc='1F1F1F',
       bg: str = None, h='left', v='center', wrap=False,
       bdr=True, fmt: str = None):
    """Write and style one cell; returns the cell."""
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    c.font      = _font(bold=bold, italic=italic, size=sz, color=fc)
    c.alignment = _al(h=h, v=v, wrap=wrap)
    if bg:
        c.fill = _fill(bg)
    if bdr:
        c.border = _border_thin()
    if fmt:
        c.number_format = fmt
    return c


def _title(ws, row: int, text: str, ncols=12, sz=14):
    """Full-width title bar (dark navy)."""
    _w(ws, row, 1, text, bold=True, sz=sz, fc=C_WHITE, bg=C_TITLE,
       h='center', bdr=False)
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row,   end_column=ncols)
    ws.row_dimensions[row].height = 28


def _sect(ws, row: int, text: str, ncols=12):
    """Section banner (medium blue)."""
    _w(ws, row, 1, text, bold=True, fc=C_WHITE, bg=C_SECT, bdr=False)
    for col in range(2, ncols + 1):
        ws.cell(row=row, column=col).fill = _fill(C_SECT)
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row,   end_column=ncols)
    ws.row_dimensions[row].height = 20


def _col_headers(ws, row: int, labels: list, start_col=1):
    """Row of bold column headers on light-blue background."""
    for i, lbl in enumerate(labels):
        _w(ws, row, start_col + i, lbl, bold=True,
           bg=C_COLHDR, h='center', bdr=True)
    ws.row_dimensions[row].height = 18


def _note(ws, row: int, text: str, ncols=12):
    """Italic grey note spanning ncols."""
    c = ws.cell(row=row, column=1)
    c.value = text
    c.font  = _font(italic=True, size=9, color='808080')
    c.alignment = _al(wrap=True)
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row,   end_column=ncols)
    ws.row_dimensions[row].height = 15


# ═══════════════════════════════════════════════════════════════════════════════
# SOURCE-DATA READING
# ═══════════════════════════════════════════════════════════════════════════════

def _read_legacy(wb) -> dict:
    """Read data from the old Book.xlsx / Blad1 layout."""
    ws = wb['Blad1']
    setup = {k: float(ws[cell].value or d)
             for k, cell, d in [('hours','B22',7.2),('shifts','B23',3),
                                 ('days','B24',6),('weeks','B25',48)]}
    demand = {}
    for r in range(3, 20):
        yr = ws.cell(row=r, column=2).value
        if yr is None:
            continue
        demand[int(yr)] = [int(ws.cell(row=r, column=c).value or 0)
                           for c in range(3, 13)]
    ct  = [[ws.cell(row=31+l, column=3+p).value or 12
            for p in range(10)] for l in range(15)]
    oee = [[ws.cell(row=49+l, column=3+p).value or 0.85
            for p in range(10)] for l in range(15)]
    mt  = [[ws.cell(row=70+i, column=3+j).value or 0
            for j in range(10)] for i in range(10)]
    ot  = [[ws.cell(row=84+i, column=3+j).value or 0
            for j in range(10)] for i in range(10)]
    return {'setup': setup, 'demand': demand, 'ct': ct, 'oee': oee,
            'mt': mt, 'ot': ot}


def _defaults() -> dict:
    demand = {yr: [0] * 10 for yr in YEARS}
    for yr, d in [
        (2029, [1000,    1000,    1000,    1000,    0,0,0,0,0,0]),
        (2030, [35498,   141992,  35498,   35498,   0,0,0,0,0,0]),
        (2031, [372560,  1490238, 372560,  372560,  0,0,0,0,0,0]),
        (2032, [529763,  2119050, 529763,  529763,  0,0,0,0,0,0]),
        (2033, [693177,  2772708, 693177,  693177,  0,0,0,0,0,0]),
        (2034, [1001694, 4006776, 1001694, 1001694, 0,0,0,0,0,0]),
        (2035, [1338697, 5354788, 1338697, 1338697, 0,0,0,0,0,0]),
        (2036, [1286832, 5417328, 1286832, 1286832, 0,0,0,0,0,0]),
        (2037, [610237,  2440948, 610237,  610237,  0,0,0,0,0,0]),
        (2038, [173752,  695008,  173752,  173752,  0,0,0,0,0,0]),
    ]:
        demand[yr] = d
    return {
        'setup':  {'hours': 7.2, 'shifts': 3, 'days': 6, 'weeks': 48},
        'demand': demand,
        'ct':  [[12.0] * 10 for _ in range(15)],
        'oee': [[0.85] * 10 for _ in range(15)],
        'mt':  [[1 if i == j else 0 for j in range(10)] for i in range(10)],
        'ot':  [[1 if i == j else 0 for j in range(10)] for i in range(10)],
    }


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET: Instructions
# ═══════════════════════════════════════════════════════════════════════════════

def _build_instructions(wb):
    ws = wb.create_sheet('Instructions')
    ws.sheet_properties.tabColor = '4472C4'
    ws.sheet_view.showGridLines  = False
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 58

    _title(ws, 1, 'LineToolingOptimization  —  ILP Solver', ncols=3, sz=15)

    ws.cell(row=2, column=2).value = \
        'Production line & tooling optimiser for multi-product manufacturing'
    ws.cell(row=2, column=2).font = _font(italic=True, size=10, color='595959')
    ws.merge_cells('B2:C2')
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8

    sections = [
        (4, 'OVERVIEW', [
            ('Goal',
             'Minimise total tooling sets (mechanical + optical) while satisfying '
             'annual demand, respecting time-based line capacity, and honouring the '
             'Line 1 validation constraint (must run all active products).'),
            ('Algorithm',
             'Mixed-Integer Linear Program solved with PuLP/CBC. All years are '
             'optimised simultaneously — avoids the myopic decisions of the legacy '
             'greedy solver (38 sets → target ≤ 28 sets).'),
            ('Priority',
             '1. Minimise lines open  ·  2. Minimise tooling sets  ·  '
             '3. Minimise product-line switches year-over-year.'),
        ]),
        (9, 'INPUT SHEETS  (edit yellow cells)', [
            ('Demand',
             'Annual unit demand per product, 2025-2041. P5-P10 are reserved for '
             'future products; set to zero until needed.'),
            ('Parameters',
             'Production setup (hours/shift, shifts, days, weeks), cycle times '
             '(seconds/unit, default 12 s), and base OEE per line (default 0.85). '
             'Changeover OEE penalty is applied by the solver in code.'),
            ('Tooling',
             'Compatibility matrices (10×10, mech + optical). Set to 1 when two '
             'products share one physical set on a line; 0 = separate sets. '
             'Currently identity — update to reflect real product families.'),
        ]),
        (14, 'OUTPUT SHEETS  (written by solver)', [
            ('Allocation',
             'One row per active (year, line) combination: product mix, units, '
             'line utilisation %, and effective OEE % applied.'),
            ('Report',
             'Lines summary, physical tooling ID registry (MECH-P01 / OPTI-P01 '
             'with line assignment and active years), demand validation.'),
        ]),
        (18, 'HOW TO RUN', [
            ('1 — Install',  'pip install pulp matplotlib openpyxl'),
            ('2 — Edit',
             'Update yellow cells in Demand, Parameters, and Tooling sheets.'),
            ('3 — Solve',    'python solver_ilp.py Solver.xlsx'),
            ('4 — Review',
             'Check Allocation and Report sheets.  '
             'Open tooling_summary.csv and line_gantt.png for additional output.'),
        ]),
    ]

    row = 4
    for sect_row, sect_title, items in sections:
        _sect(ws, sect_row, f'  {sect_title}', ncols=2)
        for label, desc in items:
            r = sect_row + 1 + items.index((label, desc))
            ws.cell(row=r, column=2).value = label
            ws.cell(row=r, column=2).font  = _font(bold=True, size=10)
            ws.cell(row=r, column=2).fill  = _fill(C_RONLY)
            ws.cell(row=r, column=2).border = _border_thin()
            ws.cell(row=r, column=3).value = desc
            ws.cell(row=r, column=3).font  = _font(size=10)
            ws.cell(row=r, column=3).alignment = _al(wrap=True)
            ws.cell(row=r, column=3).border = _border_thin()
            ws.row_dimensions[r].height = 32

    # Colour legend
    ws.row_dimensions[23].height = 8
    _sect(ws, 24, '  COLOUR CODING', ncols=2)
    for i, (color, name, desc) in enumerate([
        (C_INPUT,  'Yellow', 'Editable input cell'),
        (C_RONLY,  'Grey',   'Read-only or formula cell — do not edit'),
        (C_OUTPUT, 'Green',  'Solver output — overwritten on each run'),
        (C_COLHDR, 'Blue',   'Column header'),
    ]):
        r = 25 + i
        ws.cell(row=r, column=2).value  = name
        ws.cell(row=r, column=2).font   = _font(bold=True, size=10)
        ws.cell(row=r, column=2).fill   = _fill(color)
        ws.cell(row=r, column=2).border = _border_thin()
        ws.cell(row=r, column=3).value  = desc
        ws.cell(row=r, column=3).font   = _font(size=10)
        ws.cell(row=r, column=3).border = _border_thin()
        ws.row_dimensions[r].height = 18


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET: Demand
# ═══════════════════════════════════════════════════════════════════════════════

def _build_demand(wb, data):
    ws = wb.create_sheet('Demand')
    ws.sheet_properties.tabColor = '70AD47'
    ws.sheet_view.showGridLines  = False
    ws.freeze_panes = 'B5'

    ws.column_dimensions['A'].width = 8
    for col in range(2, 12):
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.column_dimensions[get_column_letter(12)].width = 15   # Total

    _title(ws, 1, 'VOLUME DEMAND', ncols=12)
    _note(ws, 2, 'Annual unit demand per product  ·  Edit yellow cells  ·  '
          'P5-P10 reserved for future products (leave 0 until needed)', ncols=12)
    ws.row_dimensions[3].height = 6
    _col_headers(ws, DEM_HDR_ROW, ['Year'] + PRODUCTS + ['Total'])

    demand = data['demand']
    for idx, yr in enumerate(YEARS):
        row = DEM_DATA_ROW + idx
        d   = demand.get(yr, [0] * 10)
        _w(ws, row, 1, yr, bold=True, bg=C_RONLY, h='center')

        for p in range(10):
            val = d[p] if d[p] != 0 else None
            _w(ws, row, 2 + p, val, bg=C_INPUT, h='right', fmt='#,##0')

        # Total = SUM(B:K) for this row
        first = get_column_letter(2)
        last  = get_column_letter(11)
        _w(ws, row, 12, f'=SUM({first}{row}:{last}{row})',
           bg=C_RONLY, h='right', fmt='#,##0')

        ws.row_dimensions[row].height = 16

    _note(ws, DEM_DATA_ROW + len(YEARS) + 1,
          'Demand of 0 means the product is inactive that year — the solver ignores it.',
          ncols=12)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET: Parameters
# ═══════════════════════════════════════════════════════════════════════════════

def _build_parameters(wb, data):
    ws = wb.create_sheet('Parameters')
    ws.sheet_properties.tabColor = 'ED7D31'
    ws.sheet_view.showGridLines  = False
    ws.freeze_panes = 'C14'

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 12   # units label / P1
    for col in range(3, 13):
        ws.column_dimensions[get_column_letter(col)].width = 10

    _title(ws, 1, 'PRODUCTION PARAMETERS', ncols=12)
    ws.row_dimensions[2].height = 6

    # ── Setup ──────────────────────────────────────────────────────────────────
    _sect(ws, 3, '  PRODUCTION SETUP', ncols=3)
    s = data['setup']
    setup_rows = [
        (PAR_HOURS_ROW,  'Hours per shift',          s['hours'],  '0.0',  'hrs',     True),
        (PAR_SHIFTS_ROW, 'Shifts per day',            s['shifts'], '0',    'shifts',  True),
        (PAR_DAYS_ROW,   'Working days per week',     s['days'],   '0',    'days',    True),
        (PAR_WEEKS_ROW,  'Working weeks per year',    s['weeks'],  '0',    'weeks',   True),
        (8,              'Available seconds / year',  '=B4*B5*B6*B7*3600', '#,##0', 'sec/yr', False),
        (9,              'Available hours / year',    '=B8/3600',  '#,##0', 'hrs/yr', False),
    ]
    for row, label, val, fmt, unit, editable in setup_rows:
        ws.cell(row=row, column=1).value     = label
        ws.cell(row=row, column=1).font      = _font(size=10)
        ws.cell(row=row, column=1).fill      = _fill(C_RONLY)
        ws.cell(row=row, column=1).border    = _border_thin()
        ws.cell(row=row, column=1).alignment = _al()
        _w(ws, row, 2, val, bold=editable,
           bg=C_INPUT if editable else C_RONLY, h='right', fmt=fmt)
        ws.cell(row=row, column=3).value     = unit
        ws.cell(row=row, column=3).font      = _font(size=9, color='808080')
        ws.row_dimensions[row].height        = 18

    ws.row_dimensions[10].height = 8

    # ── Cycle Time ─────────────────────────────────────────────────────────────
    _sect(ws, 11, '  CYCLE TIME  (seconds per unit)', ncols=12)
    _note(ws, 12, 'Default 12 s — adjust per product / line as needed.  '
          'Line 1 is the validation line and must support all active products.',
          ncols=12)
    _col_headers(ws, PAR_CT_HDR_ROW, ['Line'] + PRODUCTS)

    ct = data['ct']
    for l in range(15):
        row = PAR_CT_DATA_ROW + l
        _w(ws, row, 1, LINES[l], bold=True, bg=C_RONLY, h='center')
        for p in range(10):
            _w(ws, row, 2 + p, ct[l][p], bg=C_INPUT, h='center', fmt='0.0')
        ws.row_dimensions[row].height = 16

    ws.row_dimensions[PAR_CT_DATA_ROW + 15].height = 8

    # ── OEE ────────────────────────────────────────────────────────────────────
    _sect(ws, 30, '  OEE  —  Overall Equipment Effectiveness  (0.00 – 1.00)',
          ncols=12)
    _note(ws, 31,
          'Default 0.85.  The solver applies an additional −3 % penalty '
          '(configurable in solver_ilp.py) when a line runs more than one product.',
          ncols=12)
    _col_headers(ws, PAR_OEE_HDR_ROW, ['Line'] + PRODUCTS)

    oee = data['oee']
    for l in range(15):
        row = PAR_OEE_DATA_ROW + l
        _w(ws, row, 1, LINES[l], bold=True, bg=C_RONLY, h='center')
        for p in range(10):
            _w(ws, row, 2 + p, oee[l][p], bg=C_INPUT, h='center', fmt='0.00')
        ws.row_dimensions[row].height = 16


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET: Tooling
# ═══════════════════════════════════════════════════════════════════════════════

def _build_tooling(wb, data):
    ws = wb.create_sheet('Tooling')
    ws.sheet_properties.tabColor = 'FFC000'
    ws.sheet_view.showGridLines  = False

    ws.column_dimensions['A'].width = 9
    for col in range(2, 12):
        ws.column_dimensions[get_column_letter(col)].width = 8

    _title(ws, 1, 'TOOLING COMPATIBILITY MATRICES', ncols=11)
    _note(ws, 2,
          '1 = the two products share one physical tooling set on a line  '
          '(one purchase covers both)   ·   '
          '0 = separate sets required.  Matrix must be symmetric.',
          ncols=11)
    ws.row_dimensions[3].height = 8

    def _matrix(section_row, hdr_row, data_row, matrix, title):
        _sect(ws, section_row, f'  {title}', ncols=11)
        _note(ws, section_row + 1,
              'Diagonal (grey) is always 1.  Edit off-diagonal cells only.',
              ncols=11)
        _col_headers(ws, hdr_row, [''] + PRODUCTS)
        for i in range(10):
            row = data_row + i
            _w(ws, row, 1, PRODUCTS[i], bold=True, bg=C_RONLY, h='center')
            for j in range(10):
                v  = matrix[i][j]
                if i == j:
                    bg = C_RONLY        # diagonal — read-only
                elif v == 1:
                    bg = 'C6EFCE'       # sharing enabled — green tint
                else:
                    bg = C_INPUT        # editable
                _w(ws, row, 2 + j, v, bg=bg, h='center')
            ws.row_dimensions[row].height = 16

    _matrix(4, TOO_MECH_HDR_ROW, TOO_MECH_DATA_ROW, data['mt'],
            'MECHANICAL TOOLING')
    ws.row_dimensions[TOO_MECH_DATA_ROW + 10].height = 8
    _matrix(18, TOO_OPT_HDR_ROW, TOO_OPT_DATA_ROW, data['ot'],
            'OPTICAL TOOLING')

    _note(ws, TOO_OPT_DATA_ROW + 11,
          'Example: set mech[P1][P3] = mech[P3][P1] = 1 if P1 and P3 share '
          'a mechanical set — the solver will count them as one set per line.',
          ncols=11)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET: Allocation  (shell — solver fills data rows)
# ═══════════════════════════════════════════════════════════════════════════════

def _build_allocation_shell(wb):
    ws = wb.create_sheet('Allocation')
    ws.sheet_properties.tabColor = '9DC3E6'
    ws.sheet_view.showGridLines  = False
    ws.freeze_panes = f'A{ALLOC_DATA_ROW}'

    widths = [7, 7, 8, 22] + [12]*10 + [14, 10, 9]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _title(ws, 1, 'ALLOCATION RESULTS', ncols=17)

    ws.cell(row=2, column=1).value = \
        'Run  python solver_ilp.py Solver.xlsx  to populate this sheet.'
    ws.cell(row=2, column=1).font  = _font(italic=True, size=10, color='808080')
    ws.merge_cells('A2:Q2')
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6

    _sect(ws, 4, '  ALLOCATION TABLE  —  one row per active (year, line) pair',
          ncols=17)
    _col_headers(ws, ALLOC_HDR_ROW,
                 ['Year', 'Line', 'Intro', 'Products'] + PRODUCTS +
                 ['Total', 'Util %', 'OEE %'])

    # Green placeholder rows so the output area is visible before first run
    for row in range(ALLOC_DATA_ROW, ALLOC_DATA_ROW + 30):
        for col in range(1, 18):
            ws.cell(row=row, column=col).fill   = _fill(C_OUTPUT)
            ws.cell(row=row, column=col).border = _border_thin()
        ws.row_dimensions[row].height = 15

    _note(ws, ALLOC_DATA_ROW + 31,
          '↑  Solver output — overwritten on each run.  '
          'Rows below the last active year are cleared automatically.',
          ncols=17)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET: Report  (shell — solver fills data)
# ═══════════════════════════════════════════════════════════════════════════════

def _build_report_shell(wb):
    ws = wb.create_sheet('Report')
    ws.sheet_properties.tabColor = 'A9D18E'
    ws.sheet_view.showGridLines  = False

    for col, w in zip('ABCDEFG', [10, 12, 12, 30, 16, 16, 20]):
        ws.column_dimensions[col].width = w

    _title(ws, 1, 'SOLVER REPORT', ncols=7)
    ws.cell(row=2, column=1).value = \
        'Run  python solver_ilp.py Solver.xlsx  to populate this sheet.'
    ws.cell(row=2, column=1).font  = _font(italic=True, size=10, color='808080')
    ws.merge_cells('A2:G2')
    ws.row_dimensions[2].height = 18

    for section_row, title in [(4, 'KEY METRICS'),
                                (9, 'LINES SUMMARY'),
                                (20, 'TOOLING ID REGISTRY'),
                                (36, 'DEMAND VALIDATION')]:
        _sect(ws, section_row, f'  {section_row and title}', ncols=7)
        for r in range(section_row + 1, section_row + 9):
            for col in range(1, 8):
                ws.cell(row=r, column=col).fill   = _fill(C_OUTPUT)
                ws.cell(row=r, column=col).border = _border_thin()
            ws.row_dimensions[r].height = 15
        ws.cell(row=section_row + 1, column=1).value = \
            '↑ Solver output'
        ws.cell(row=section_row + 1, column=1).font = \
            _font(italic=True, size=9, color='808080')
        ws.merge_cells(start_row=section_row + 1, start_column=1,
                       end_row=section_row + 1,   end_column=7)


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description='Create Solver.xlsx for LineToolingOptimization')
    parser.add_argument('source', nargs='?', default=None,
                        help='Source workbook to copy data from (e.g. Book.xlsx)')
    parser.add_argument('-o', '--output', default='Solver.xlsx')
    args = parser.parse_args()

    out_path = Path(args.output)

    if args.source:
        src = Path(args.source)
        if not src.exists():
            print(f'Error: {src} not found'); sys.exit(1)
        print(f'Reading data from {src} ...')
        src_wb = openpyxl.load_workbook(str(src), data_only=True)
        data   = _read_legacy(src_wb)
    else:
        # Try Book.xlsx in the same directory automatically
        auto = Path(__file__).parent / 'Book.xlsx'
        if auto.exists():
            print(f'Auto-detected {auto} — reading data ...')
            src_wb = openpyxl.load_workbook(str(auto), data_only=True)
            data   = _read_legacy(src_wb)
        else:
            print('No source workbook found — using built-in defaults.')
            data = _defaults()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # delete the default blank sheet

    print('Building sheets ...')
    _build_instructions(wb)
    _build_demand(wb, data)
    _build_parameters(wb, data)
    _build_tooling(wb, data)
    _build_allocation_shell(wb)
    _build_report_shell(wb)

    wb.save(str(out_path))
    print(f'Saved → {out_path}')
    print()
    print('Next steps:')
    print('  1. Open Solver.xlsx and review / adjust yellow input cells.')
    print('  2. pip install pulp matplotlib openpyxl')
    print('  3. python solver_ilp.py Solver.xlsx')


if __name__ == '__main__':
    main()

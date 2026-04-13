#!/usr/bin/env python3
"""
solver_simple.py — Simplified ILP Solver for LineToolingOptimization
=====================================================================
A stripped-down alternative to solver_ilp.py.  No tooling tracking,
no late-introduction costs.  Three rules drive the model:

  1. Line 1 runs all active products at least once (validation).
  2. New lines open only when existing capacity is exhausted.
  3. Prefer single-product lines (fewer changeovers) via a soft penalty.

Objective:
    minimise  running_cost × Σ o[l,i]
            + changeover_penalty × Σ mul[l,i]

This is significantly smaller than solver_ilp.py (~3 000 vs ~5 300 variables)
and typically solves in under 60 seconds.

Usage:
    python solver_simple.py Simple.xlsx
    python solver_simple.py            # looks for Simple.xlsx in current dir
Requirements: pip install pulp openpyxl
"""
import sys
import argparse
import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell
import pulp

# ── Sheet names ───────────────────────────────────────────────────────────────
DEMAND_SHEET = 'Demand'
PARAMS_SHEET = 'Parameters'
ALLOC_SHEET  = 'Allocation'

NUM_PRODUCTS = 10
NUM_LINES    = 15

# 'Demand' sheet positions (must match create_simple_xlsx.py)
DEM_DATA_ROW = 5

# 'Parameters' sheet positions
PAR_HOURS_ROW     = 4
PAR_SHIFTS_ROW    = 5
PAR_DAYS_ROW      = 6
PAR_WEEKS_ROW     = 7
PAR_VAL_COL       = 2
PAR_CT_DATA_ROW   = 14
PAR_OEE_DATA_ROW  = 34   # first OEE data row (row 31=sect, 32=note, 33=col-hdr, 34=L1)
PAR_COST_RUN_ROW     = 51
PAR_COST_CHGOVER_ROW = 52

# 'Allocation' sheet output positions
ALLOC_HDR_ROW  = 5
ALLOC_DATA_ROW = 6

# ── Solver configuration ──────────────────────────────────────────────────────
BASE_OEE               = 0.85
CHANGEOVER_OEE_PENALTY = 0.03
SOLVER_TIME_LIMIT      = 300   # seconds

# Minimum raw utilisation a line 2+ must achieve in its opening year.
# Blocks commissioning a new line before existing lines are at capacity.
MIN_OPENING_UTIL = 0.30

# Default costs (overridden by Parameters sheet)
DEFAULT_COST_RUNNING    = 500_000   # USD / open line / year
DEFAULT_COST_CHANGEOVER = 100_000   # USD / mixed-product line-year

# Colour palette (matches create_simple_xlsx.py)
C_TITLE  = '1F3864'
C_SECT   = '2E75B6'
C_COLHDR = 'BDD7EE'
C_RONLY  = 'F2F2F2'
C_GREEN  = 'E2EFDA'   # single-product line cell
C_AMBER  = 'FFE699'   # mixed-product line cell
C_WHITE  = 'FFFFFF'
C_BORDER = 'B8CCE4'


# ─────────────────────────────────────────────────────────────────────────────
# INPUT LOADING
# ─────────────────────────────────────────────────────────────────────────────
def load_inputs(wb):
    ws_d = wb[DEMAND_SHEET]
    ws_p = wb[PARAMS_SHEET]

    setup = {
        'hours':  float(ws_p.cell(row=PAR_HOURS_ROW,  column=PAR_VAL_COL).value or 0),
        'shifts': float(ws_p.cell(row=PAR_SHIFTS_ROW, column=PAR_VAL_COL).value or 0),
        'days':   float(ws_p.cell(row=PAR_DAYS_ROW,   column=PAR_VAL_COL).value or 0),
        'weeks':  float(ws_p.cell(row=PAR_WEEKS_ROW,  column=PAR_VAL_COL).value or 0),
    }
    avail = setup['hours'] * setup['shifts'] * setup['days'] * setup['weeks'] * 3600

    years, demand = [], {}
    for r in range(DEM_DATA_ROW, DEM_DATA_ROW + 50):
        yr = ws_d.cell(row=r, column=1).value
        if yr is None:
            break
        yr = int(yr)
        years.append(yr)
        demand[yr] = [int(ws_d.cell(row=r, column=2 + p).value or 0)
                      for p in range(10)]

    ct  = [[ws_p.cell(row=PAR_CT_DATA_ROW  + l, column=2 + p).value or 12
            for p in range(10)] for l in range(15)]
    # OEE: Simple.xlsx layout has the OEE col-header at row 34 and data from 35
    oee = [[ws_p.cell(row=PAR_OEE_DATA_ROW + l, column=2 + p).value or 0.85
            for p in range(10)] for l in range(15)]

    def _cost(row, default):
        v = ws_p.cell(row=row, column=PAR_VAL_COL).value
        return float(v) if v is not None else default

    costs = {
        'running':    _cost(PAR_COST_RUN_ROW,     DEFAULT_COST_RUNNING),
        'changeover': _cost(PAR_COST_CHGOVER_ROW, DEFAULT_COST_CHANGEOVER),
    }

    return {'years': years, 'demand': demand, 'avail': avail,
            'days': setup['days'],
            'ct': ct, 'oee': oee, 'costs': costs}


# ─────────────────────────────────────────────────────────────────────────────
# ILP BUILD & SOLVE
# ─────────────────────────────────────────────────────────────────────────────
def build_and_solve(inp):
    """
    Build and solve the simplified MIP.

    Returns (alloc, intro, n_line_years, n_mixed_years) or None on failure.
      alloc  : dict (year, line, product) -> int units
      intro  : dict  line -> first year the line produced
    """
    demand = inp['demand']
    avail  = inp['avail']
    ct     = inp['ct']
    costs  = inp['costs']

    # Only model years with actual demand
    years = [yr for yr in inp['years'] if sum(demand[yr]) > 0]

    P = NUM_PRODUCTS
    L = NUM_LINES
    Y = len(years)

    M_big = max(demand[yr][p] for yr in years for p in range(P)) + 1

    prob = pulp.LpProblem('SimpleLineTooling', pulp.LpMinimize)

    # ── Decision variables ────────────────────────────────────────────────────
    # x[p,l,i]  continuous: units of product p on line l in year index i
    x = pulp.LpVariable.dicts(
        'x',
        [(p, l, i) for p in range(P) for l in range(L) for i in range(Y)],
        lowBound=0, cat='Continuous')

    # u[p,l,i]  binary: 1 if product p is produced on line l in year i
    u = pulp.LpVariable.dicts(
        'u',
        [(p, l, i) for p in range(P) for l in range(L) for i in range(Y)],
        cat='Binary')

    # o[l,i]  binary: 1 if line l is open in year i
    o = pulp.LpVariable.dicts(
        'open',
        [(l, i) for l in range(L) for i in range(Y)],
        cat='Binary')

    # mul[l,i]  binary: 1 if line l runs 2+ products in year i
    mul = pulp.LpVariable.dicts(
        'multi',
        [(l, i) for l in range(L) for i in range(Y)],
        cat='Binary')

    # ── Objective ─────────────────────────────────────────────────────────────
    prob += (
        costs['running']    * pulp.lpSum(o[l, i]   for l in range(L) for i in range(Y))
      + costs['changeover'] * pulp.lpSum(mul[l, i] for l in range(L) for i in range(Y))
    )

    # ── Constraints ───────────────────────────────────────────────────────────
    for i, yr in enumerate(years):
        d = demand[yr]

        # 1. Demand satisfaction
        for p in range(P):
            prob += (pulp.lpSum(x[p, l, i] for l in range(L)) == d[p],
                     f'dem_{p}_{i}')

        # 2. Zero-demand products: no production or assignment flags
        for p in range(P):
            if d[p] == 0:
                for l in range(L):
                    prob += (u[p, l, i] == 0, f'zero_{p}_{l}_{i}')

        for l in range(L):
            # 3. x <= M * u  (production only when flagged)
            for p in range(P):
                prob += (x[p, l, i] <= M_big * u[p, l, i], f'xu_{p}_{l}_{i}')

            # 4. Line open iff at least one product assigned
            for p in range(P):
                prob += (o[l, i] >= u[p, l, i], f'open_ge_{p}_{l}_{i}')
            prob += (o[l, i] <= pulp.lpSum(u[p, l, i] for p in range(P)),
                     f'open_le_{l}_{i}')

            # 5. Multi-product flag: forced to 1 when 2+ products present
            prob += (pulp.lpSum(u[p, l, i] for p in range(P)) - 1
                     <= (P - 1) * mul[l, i],
                     f'multi_{l}_{i}')
            prob += (mul[l, i] <= o[l, i], f'multi_le_open_{l}_{i}')

            # 6. Capacity with OEE penalty
            #    sum(x*ct) + avail * PENALTY * mul <= avail * BASE_OEE
            prob += (
                pulp.lpSum(x[p, l, i] * ct[l][p] for p in range(P))
                + avail * CHANGEOVER_OEE_PENALTY * mul[l, i]
                <= avail * BASE_OEE,
                f'cap_{l}_{i}'
            )

    # 7. Line 1 validation: each active product must run on Line 1 at least once
    for p in range(P):
        if any(demand[yr][p] > 0 for yr in years):
            prob += (
                pulp.lpSum(u[p, 0, i] for i in range(Y)) >= 1,
                f'val1_{p}'
            )

    # 8a. Lines don't close: once open, stays open
    for l in range(L):
        for i in range(1, Y):
            prob += (o[l, i] >= o[l, i - 1], f'no_close_{l}_{i}')

    # 8b. Minimum opening-year utilisation for lines 2+
    #     A new line (o transitions 0→1) must reach MIN_OPENING_UTIL of raw capacity.
    for l in range(1, L):
        for i in range(1, Y):
            prob += (
                pulp.lpSum(x[p, l, i] * ct[l][p] for p in range(P))
                >= MIN_OPENING_UTIL * avail * (o[l, i] - o[l, i - 1]),
                f'min_open_util_{l}_{i}'
            )

    # ── Solve ─────────────────────────────────────────────────────────────────
    n_vars = len(prob.variables())
    n_cons = len(prob.constraints)
    print(f'  Variables: {n_vars:,}   Constraints: {n_cons:,}')

    solver = pulp.PULP_CBC_CMD(timeLimit=SOLVER_TIME_LIMIT, msg=1)
    prob.solve(solver)

    status  = pulp.LpStatus[prob.status]
    obj_val = pulp.value(prob.objective)
    print(f'  Solver status: {status}   Objective: {obj_val:,.0f}')

    if prob.status not in (1, -2):
        print('  ERROR: solver did not find a feasible solution.')
        return None

    # ── Extract solution ───────────────────────────────────────────────────────
    alloc = {}
    intro = {}

    for i, yr in enumerate(years):
        for l in range(L):
            for p in range(P):
                v = pulp.value(x[p, l, i])
                if v is not None and v > 0.5:
                    alloc[(yr, l, p)] = round(v)
                    if l not in intro:
                        intro[l] = yr

    n_line_years  = sum(1 for l in range(L) for i in range(Y)
                        if pulp.value(o[l, i]) is not None
                        and pulp.value(o[l, i]) > 0.5)
    n_mixed_years = sum(1 for l in range(L) for i in range(Y)
                        if pulp.value(mul[l, i]) is not None
                        and pulp.value(mul[l, i]) > 0.5)

    return alloc, intro, n_line_years, n_mixed_years


# ─────────────────────────────────────────────────────────────────────────────
# VERIFICATION
# ─────────────────────────────────────────────────────────────────────────────
def verify(alloc, inp):
    ok = True
    years = [yr for yr in inp['years'] if sum(inp['demand'][yr]) > 0]
    for yr in years:
        d = inp['demand'][yr]

        # Demand check
        ta = [0] * NUM_PRODUCTS
        for l in range(NUM_LINES):
            for p in range(NUM_PRODUCTS):
                ta[p] += alloc.get((yr, l, p), 0)
        for p in range(NUM_PRODUCTS):
            if abs(d[p] - ta[p]) > 1:
                print(f'  X {yr} P{p+1}: demand={d[p]:,}  alloc={ta[p]:,}')
                ok = False

        # Capacity check
        for l in range(NUM_LINES):
            units_l = [alloc.get((yr, l, p), 0) for p in range(NUM_PRODUCTS)]
            t = sum(units_l[p] * inp['ct'][l][p] for p in range(NUM_PRODUCTS))
            n_prod = sum(1 for p in range(NUM_PRODUCTS) if units_l[p] > 0)
            oee_eff = BASE_OEE - (CHANGEOVER_OEE_PENALTY if n_prod > 1 else 0)
            cap = inp['avail'] * oee_eff
            if t > cap * 1.001:
                print(f'  X {yr} Line {l+1}: capacity exceeded '
                      f'({t:,.0f} > {cap:,.0f}, OEE={oee_eff*100:.0f}%)')
                ok = False
    return ok


# ─────────────────────────────────────────────────────────────────────────────
# OUTPUT HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def _ofill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def _oborder():
    thin = Side(style='thin', color=C_BORDER)
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _oal(h='center', v='center'):
    return Alignment(horizontal=h, vertical=v)


# ─────────────────────────────────────────────────────────────────────────────
# 7-DAY WORKING CHECK  (written below the main allocation grid)
# ─────────────────────────────────────────────────────────────────────────────
def _write_7day_check(ws, alloc, intro, inp, used_lines, years, row_start):
    """
    Writes a what-if section: 'If we had not opened the last line, how many of
    the remaining lines would need to work 7 days/week to cover the demand?'

    Red cells = line must extend to 7-day working.
    Returns next free row number.
    """
    P       = NUM_PRODUCTS
    avail_6 = inp['avail']
    days_6  = inp['days']
    avail_7 = avail_6 / days_6 * 7
    ct      = inp['ct']
    demand  = inp['demand']

    if len(intro) < 2:
        return row_start

    last_line  = max(intro, key=lambda l: (intro[l], l))
    last_intro = intro[last_line]
    n1_lines   = [l for l in used_lines if l != last_line]
    N1         = len(n1_lines)

    years_check = [yr for yr in years if yr >= last_intro]
    if not years_check:
        return row_start

    # ── Per-year analysis ──────────────────────────────────────────────────────
    year_meta   = {}
    any_unavoid = False

    for yr in years_check:
        d       = demand[yr]
        tot_dem = sum(d)

        shortfall_sec = sum(
            alloc.get((yr, last_line, p), 0) * ct[last_line][p]
            for p in range(P)
        )

        line_extra = {}
        for l in n1_lines:
            units_l = [alloc.get((yr, l, p), 0) for p in range(P)]
            n_prod  = sum(1 for p in range(P) if units_l[p] > 0)
            oee_eff = BASE_OEE - (CHANGEOVER_OEE_PENALTY if n_prod > 1 else 0)
            line_extra[l] = (avail_7 - avail_6) * oee_eff

        total_extra = sum(line_extra.values())
        unavoidable = shortfall_sec > 0 and total_extra < shortfall_sec

        if unavoidable:
            any_unavoid = True
            lines_7day  = set(n1_lines)
        elif shortfall_sec <= 0:
            lines_7day = set()
        else:
            sorted_l   = sorted(n1_lines, key=lambda l: line_extra[l], reverse=True)
            lines_7day = set()
            covered    = 0.0
            for l in sorted_l:
                if covered >= shortfall_sec:
                    break
                lines_7day.add(l)
                covered += line_extra[l]

        tot_cap = 0
        for l in n1_lines:
            units_l = [alloc.get((yr, l, p), 0) for p in range(P)]
            n_prod  = sum(1 for p in range(P) if units_l[p] > 0)
            oee_eff = BASE_OEE - (CHANGEOVER_OEE_PENALTY if n_prod > 1 else 0)
            av      = avail_7 if l in lines_7day else avail_6
            tot_ct  = sum(ct[l][p] * d[p] for p in range(P))
            avg_ct  = (tot_ct / tot_dem) if tot_dem > 0 else 12
            tot_cap += av * oee_eff / avg_ct

        year_meta[yr] = {
            'lines_7day':  lines_7day,
            'unavoidable': unavoidable,
            'tot_cap':     tot_cap,
        }

    # ── Write to sheet ────────────────────────────────────────────────────────
    C_RED    = 'FFC7CE'
    C_RED_FT = '9C0006'
    bdr      = _oborder()
    row      = row_start + 2

    # Banner
    ws.cell(row=row, column=1).value = (
        f'7-DAY WORKING CHECK  —  Alternative to opening L{last_line + 1}'
    )
    ws.cell(row=row, column=1).font      = Font(name='Calibri', bold=True,
                                                size=12, color=C_WHITE)
    ws.cell(row=row, column=1).fill      = PatternFill('solid', fgColor=C_TITLE)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center',
                                                      vertical='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row,
                   end_column=max(N1 + 5, 8))
    ws.row_dimensions[row].height = 24
    row += 1

    ws.cell(row=row, column=1).value = (
        f'From {last_intro} onward: could the {N1} existing line(s) have '
        f'avoided L{last_line + 1} by working 7 days/week?  '
        f'(current schedule: {int(days_6)} days/week)'
    )
    ws.cell(row=row, column=1).font = Font(name='Calibri', italic=True,
                                            size=9, color='595959')
    ws.merge_cells(start_row=row, start_column=1, end_row=row,
                   end_column=max(N1 + 5, 8))
    ws.row_dimensions[row].height = 14
    row += 1
    ws.row_dimensions[row].height = 6
    row += 1

    extra_start = 2 + N1
    ws.column_dimensions['A'].width = 8
    for idx in range(N1):
        ws.column_dimensions[get_column_letter(2 + idx)].width = 14
    for i, w in enumerate([12, 14, 14, 10]):
        ws.column_dimensions[get_column_letter(extra_start + i)].width = w

    def _h(r, c, val):
        cell = ws.cell(row=r, column=c)
        cell.value     = val
        cell.font      = Font(name='Calibri', bold=True, size=10)
        cell.fill      = PatternFill('solid', fgColor=C_COLHDR)
        cell.border    = bdr
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)

    ws.row_dimensions[row].height = 28
    _h(row, 1, 'Year')
    for idx, l in enumerate(n1_lines):
        _h(row, 2 + idx, f'L{l + 1}')
    _h(row, extra_start,     'Lines\nat 7-day')
    _h(row, extra_start + 1, 'Total\nDemand')
    _h(row, extra_start + 2, 'Capacity\n(scenario)')
    _h(row, extra_start + 3, 'Util %')
    row += 1

    for yr in years_check:
        meta     = year_meta[yr]
        lines7   = meta['lines_7day']
        unavoid  = meta['unavoidable']
        d        = demand[yr]
        tot_dem  = sum(d)
        tot_cap  = meta['tot_cap']
        util_pct = (tot_dem / tot_cap * 100) if tot_cap > 0 else 0

        ws.row_dimensions[row].height = 18

        yc = ws.cell(row=row, column=1)
        yc.value     = yr
        yc.font      = Font(name='Calibri', bold=True, size=10)
        yc.fill      = PatternFill('solid', fgColor=C_RED if unavoid else C_RONLY)
        yc.border    = bdr
        yc.alignment = Alignment(horizontal='center', vertical='center')

        for idx, l in enumerate(n1_lines):
            col  = 2 + idx
            cell = ws.cell(row=row, column=col)
            cell.border = bdr
            prods = [p for p in range(P) if alloc.get((yr, l, p), 0) > 0]
            at_7  = l in lines7

            if not prods:
                cell.fill      = PatternFill('solid', fgColor=C_RED if at_7 else 'F2F2F2')
                cell.value     = '7d —' if at_7 else '—'
                cell.font      = Font(name='Calibri', size=9, italic=True,
                                      bold=at_7,
                                      color=C_RED_FT if at_7 else 'AAAAAA')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                label = ', '.join(f'P{p+1}' for p in prods)
                multi = len(prods) > 1
                bg    = C_RED if at_7 else (C_AMBER if multi else C_GREEN)
                color = C_RED_FT if at_7 else '1F1F1F'
                cell.fill      = PatternFill('solid', fgColor=bg)
                cell.value     = ('7d  ' if at_7 else '') + label
                cell.font      = Font(name='Calibri', size=9, bold=True, color=color)
                cell.alignment = Alignment(horizontal='center', vertical='center',
                                           wrap_text=True)

        sum_bg = C_RED if unavoid else C_RONLY
        sum_ft = C_RED_FT if unavoid else '1F1F1F'

        def _sc(col, val, fmt=None):
            c = ws.cell(row=row, column=col)
            c.value     = val
            c.font      = Font(name='Calibri', size=10, color=sum_ft)
            c.fill      = PatternFill('solid', fgColor=sum_bg)
            c.border    = bdr
            c.alignment = Alignment(horizontal='right', vertical='center')
            if fmt and isinstance(val, (int, float)):
                c.number_format = fmt

        n7_label = '⚠ NOT ENOUGH' if unavoid else len(lines7)
        _sc(extra_start,     n7_label)
        _sc(extra_start + 1, tot_dem,            '#,##0')
        _sc(extra_start + 2, int(tot_cap),       '#,##0')
        _sc(extra_start + 3, round(util_pct, 1), '0.0"%"')
        row += 1

    # Legend
    row += 1
    for bg, label in [
        (C_RED,   'Line must work 7 days/week to absorb demand from removed line'),
        (C_GREEN, 'Line at normal 6 days/week — single product'),
        (C_AMBER, 'Line at normal 6 days/week — multi product'),
    ]:
        c = ws.cell(row=row, column=2)
        c.fill   = PatternFill('solid', fgColor=bg)
        c.value  = ''
        c.border = bdr
        c2 = ws.cell(row=row, column=3)
        c2.value = label
        c2.font  = Font(name='Calibri', italic=True, size=9, color='595959')
        ws.row_dimensions[row].height = 14
        row += 1

    if any_unavoid:
        row += 1
        bad_yrs = ', '.join(str(yr) for yr in years_check
                            if year_meta[yr]['unavoidable'])
        note = (
            f'\u26a0  NOTE: Even with all {N1} existing line(s) at 7 days/week, '
            f'demand in {bad_yrs} exceeds their total capacity. '
            f'Opening L{last_line + 1} (or equivalent) is unavoidable.'
        )
        c = ws.cell(row=row, column=1)
        c.value     = note
        c.font      = Font(name='Calibri', bold=True, size=10, color=C_RED_FT)
        c.fill      = PatternFill('solid', fgColor='FFE7E7')
        c.alignment = Alignment(horizontal='left', vertical='center',
                                wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row,
                       end_column=max(N1 + 5, 8))
        ws.row_dimensions[row].height = 30
        row += 1

    return row


# ─────────────────────────────────────────────────────────────────────────────
# WRITE OUTPUT  (Allocation sheet)
# ─────────────────────────────────────────────────────────────────────────────
def write_output(wb, alloc, intro, inp, n_line_years, n_mixed_years, ok):
    ws = wb[ALLOC_SHEET]
    avail = inp['avail']
    ct    = inp['ct']
    years_all  = inp['years']
    demand     = inp['demand']
    costs      = inp['costs']

    # Active demand years only
    years = [yr for yr in years_all if sum(demand[yr]) > 0]

    # Which lines were ever used?
    used_lines = sorted(intro.keys())
    N = len(used_lines)    # number of lines used

    # ── Clear old content (rows 1 onward, including title and header rows) ────
    clear_start = 1
    clear_end   = ALLOC_DATA_ROW + 200
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row <= clear_end and rng.max_row >= clear_start:
            ws.unmerge_cells(str(rng))
    for r in range(clear_start, clear_end):
        for c in range(1, max(N + 10, 20)):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            cell.value  = None
            cell.fill   = _ofill(C_WHITE)
            cell.border = _oborder()

    # ── Title & subtitle ───────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    c1 = ws.cell(row=1, column=1)
    c1.value = 'ALLOCATION  —  solver_simple output'
    c1.font  = Font(name='Calibri', bold=True, size=14, color=C_WHITE)
    c1.fill  = _ofill(C_TITLE)
    c1.alignment = Alignment(horizontal='center', vertical='center')
    if N + 5 > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1,
                       end_column=max(N + 5, 8))

    ts = datetime.datetime.now().strftime('%Y-%m-%d  %H:%M')
    status_txt = '✓  All checks passed' if ok else '⚠  Issues found'
    ws.cell(row=2, column=1).value = (
        f'Last run: {ts}   |   {status_txt}   |   '
        f'{len(used_lines)} lines used'
    )
    ws.cell(row=2, column=1).font = Font(name='Calibri', italic=True,
                                         size=9, color='595959')
    ws.merge_cells(start_row=2, start_column=1, end_row=2,
                   end_column=max(N + 5, 8))
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 6
    ws.row_dimensions[4].height = 6

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 8   # Year
    for idx, l in enumerate(used_lines):
        ws.column_dimensions[get_column_letter(2 + idx)].width = 14
    extra_start = 2 + N
    for i, w in enumerate([10, 14, 14, 12]):   # Open | Demand | Capacity | Util%
        ws.column_dimensions[get_column_letter(extra_start + i)].width = w

    # ── Header row ─────────────────────────────────────────────────────────────
    thin = Side(style='thin', color=C_BORDER)
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _hdr(row, col, val):
        c = ws.cell(row=row, column=col)
        c.value     = val
        c.font      = Font(name='Calibri', bold=True, size=10)
        c.fill      = _ofill(C_COLHDR)
        c.border    = bdr
        c.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[ALLOC_HDR_ROW].height = 18
    _hdr(ALLOC_HDR_ROW, 1, 'Year')
    for idx, l in enumerate(used_lines):
        _hdr(ALLOC_HDR_ROW, 2 + idx, f'L{l+1}')
    _hdr(ALLOC_HDR_ROW, extra_start,     'Open\nLines')
    _hdr(ALLOC_HDR_ROW, extra_start + 1, 'Total\nDemand')
    _hdr(ALLOC_HDR_ROW, extra_start + 2, 'Total\nCapacity')
    _hdr(ALLOC_HDR_ROW, extra_start + 3, 'Util %')

    # ── Data rows ──────────────────────────────────────────────────────────────
    row = ALLOC_DATA_ROW
    for yr in years:
        d        = demand[yr]
        tot_dem  = sum(d)
        n_open   = sum(1 for l in used_lines
                       if any(alloc.get((yr, l, p), 0) > 0 for p in range(NUM_PRODUCTS))
                       or (l in intro and intro[l] <= yr))

        # Capacity of all open lines (considering multi-product OEE penalty)
        tot_cap = 0
        for l in used_lines:
            if l not in intro or intro[l] > yr:
                continue
            units_l = [alloc.get((yr, l, p), 0) for p in range(NUM_PRODUCTS)]
            n_prod  = sum(1 for p in range(NUM_PRODUCTS) if units_l[p] > 0)
            oee_eff = BASE_OEE - (CHANGEOVER_OEE_PENALTY if n_prod > 1 else 0)
            tot_cap += avail * oee_eff / 12   # approx units capacity (12 s default ct)

        util_pct = (tot_dem / tot_cap * 100) if tot_cap > 0 else 0

        ws.row_dimensions[row].height = 18

        # Year cell
        c = ws.cell(row=row, column=1)
        c.value     = yr
        c.font      = Font(name='Calibri', bold=True, size=10)
        c.fill      = _ofill(C_RONLY)
        c.border    = bdr
        c.alignment = Alignment(horizontal='center', vertical='center')

        # Line cells
        for idx, l in enumerate(used_lines):
            col  = 2 + idx
            cell = ws.cell(row=row, column=col)
            cell.border = bdr

            if l not in intro or intro[l] > yr:
                cell.fill  = _ofill(C_WHITE)
                cell.value = None
                continue

            prods = [p for p in range(NUM_PRODUCTS)
                     if alloc.get((yr, l, p), 0) > 0]
            if not prods:
                # Open but idle this year
                cell.fill  = _ofill('F2F2F2')
                cell.value = '—'
                cell.font  = Font(name='Calibri', size=9, color='AAAAAA',
                                  italic=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                label = ', '.join(f'P{p+1}' for p in prods)
                multi = len(prods) > 1
                cell.fill  = _ofill(C_AMBER if multi else C_GREEN)
                cell.value = label
                cell.font  = Font(name='Calibri', size=9, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center',
                                           wrap_text=True)

        # Summary cells
        def _sum_cell(col, val, fmt=None):
            c = ws.cell(row=row, column=col)
            c.value     = val
            c.font      = Font(name='Calibri', size=10)
            c.fill      = _ofill(C_RONLY)
            c.border    = bdr
            c.alignment = Alignment(horizontal='right', vertical='center')
            if fmt:
                c.number_format = fmt

        _sum_cell(extra_start,     n_open)
        _sum_cell(extra_start + 1, tot_dem,  '#,##0')
        _sum_cell(extra_start + 2, int(tot_cap), '#,##0')
        _sum_cell(extra_start + 3, round(util_pct, 1), '0.0"%"')
        row += 1

    # ── Colour legend ──────────────────────────────────────────────────────────
    row += 1
    for color, label in [(C_GREEN, 'Single-product line (no changeover)'),
                         (C_AMBER, 'Multi-product line (changeover penalty applies)'),
                         ('F2F2F2', 'Line open but idle this year')]:
        c = ws.cell(row=row, column=2)
        c.fill  = _ofill(color)
        c.value = ''
        c.border = bdr
        c2 = ws.cell(row=row, column=3)
        c2.value = label
        c2.font  = Font(name='Calibri', italic=True, size=9, color='595959')
        ws.row_dimensions[row].height = 14
        row += 1

    # ── Summary block ──────────────────────────────────────────────────────────
    row += 1
    ws.cell(row=row, column=1).value = 'LINES USED'
    ws.cell(row=row, column=1).font  = Font(name='Calibri', bold=True, size=10,
                                            color=C_WHITE)
    ws.cell(row=row, column=1).fill  = _ofill(C_SECT)
    ws.merge_cells(start_row=row, start_column=1, end_row=row,
                   end_column=max(N + 5, 6))
    ws.row_dimensions[row].height = 18
    row += 1

    for l in used_lines:
        prods_ever = sorted(set(p for yr in years
                                for p in range(NUM_PRODUCTS)
                                if alloc.get((yr, l, p), 0) > 0))
        n_prod_ever = len(prods_ever)
        oee_eff = BASE_OEE - (CHANGEOVER_OEE_PENALTY if n_prod_ever > 1 else 0)
        label = (f'L{l+1}  intro={intro[l]}  '
                 f'products=[{", ".join(f"P{p+1}" for p in prods_ever)}]  '
                 f'OEE={oee_eff*100:.0f}%')
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font  = Font(name='Calibri', size=10)
        ws.cell(row=row, column=1).fill  = _ofill(C_RONLY)
        ws.merge_cells(start_row=row, start_column=1, end_row=row,
                       end_column=max(N + 5, 6))
        ws.row_dimensions[row].height = 16
        row += 1

    # ── Cost breakdown ─────────────────────────────────────────────────────────
    row += 1
    ws.cell(row=row, column=1).value = 'COST BREAKDOWN'
    ws.cell(row=row, column=1).font  = Font(name='Calibri', bold=True, size=10,
                                            color=C_WHITE)
    ws.cell(row=row, column=1).fill  = _ofill(C_SECT)
    ws.merge_cells(start_row=row, start_column=1, end_row=row,
                   end_column=max(N + 5, 6))
    ws.row_dimensions[row].height = 18
    row += 1

    cost_running    = costs['running']    * n_line_years
    cost_changeover = costs['changeover'] * n_mixed_years
    total_cost      = cost_running + cost_changeover

    for label, val in [
        (f'Running cost: {n_line_years} line-years × ${costs["running"]:,.0f}',
         cost_running),
        (f'Changeover penalty: {n_mixed_years} mixed-line-years × ${costs["changeover"]:,.0f}',
         cost_changeover),
        ('TOTAL', total_cost),
    ]:
        bold = (label == 'TOTAL')
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font  = Font(name='Calibri', bold=bold, size=10)
        ws.cell(row=row, column=1).fill  = _ofill(C_RONLY)
        ws.merge_cells(start_row=row, start_column=1, end_row=row,
                       end_column=max(N + 3, 5))
        c = ws.cell(row=row, column=max(N + 4, 6))
        c.value          = val
        c.font           = Font(name='Calibri', bold=bold, size=10)
        c.fill           = _ofill(C_RONLY)
        c.number_format  = '$#,##0'
        c.alignment      = Alignment(horizontal='right')
        ws.row_dimensions[row].height = 16
        row += 1

    # ── 7-day working check ────────────────────────────────────────────────────
    row = _write_7day_check(ws, alloc, intro, inp, used_lines, years, row)

    return row - ALLOC_DATA_ROW  # rows written


# ─────────────────────────────────────────────────────────────────────────────
# CONSOLE SUMMARY
# ─────────────────────────────────────────────────────────────────────────────
def print_summary(alloc, intro, inp, n_line_years, n_mixed_years):
    demand = inp['demand']
    avail  = inp['avail']
    ct     = inp['ct']
    costs  = inp['costs']
    years  = [yr for yr in inp['years'] if sum(demand[yr]) > 0]

    print('\nLines used:')
    for l in sorted(intro.keys()):
        prods = sorted(set(p for yr in years for p in range(NUM_PRODUCTS)
                           if alloc.get((yr, l, p), 0) > 0))
        n_p   = len(prods)
        oee   = BASE_OEE - (CHANGEOVER_OEE_PENALTY if n_p > 1 else 0)
        print(f'  Line {l+1:2d}  intro={intro[l]}  '
              f'products=[{", ".join(f"P{p+1}" for p in prods)}]  '
              f'OEE={oee*100:.0f}%')

    print('\nPer-year summary:')
    for yr in years:
        d = demand[yr]
        active_lines = [(l, [p for p in range(NUM_PRODUCTS)
                              if alloc.get((yr, l, p), 0) > 0])
                        for l in sorted(intro.keys())
                        if intro[l] <= yr and
                        any(alloc.get((yr, l, p), 0) > 0 for p in range(NUM_PRODUCTS))]
        parts = [f'L{l+1}=[{",".join(f"P{p+1}" for p in prods)}]'
                 for l, prods in active_lines if prods]
        n_open = sum(1 for l in intro if intro[l] <= yr)

        # utilisation = total units × cycle_time / sum_of_open_line_capacity
        tot_time = sum(alloc.get((yr, l, p), 0) * ct[l][p]
                       for l in sorted(intro.keys()) if intro[l] <= yr
                       for p in range(NUM_PRODUCTS))
        tot_cap = 0
        for l in sorted(intro.keys()):
            if intro[l] > yr:
                continue
            units_l = [alloc.get((yr, l, p), 0) for p in range(NUM_PRODUCTS)]
            n_prod  = sum(1 for p in range(NUM_PRODUCTS) if units_l[p] > 0)
            oee_eff = BASE_OEE - (CHANGEOVER_OEE_PENALTY if n_prod > 1 else 0)
            tot_cap += avail * oee_eff
        util = (tot_time / tot_cap * 100) if tot_cap > 0 else 0

        print(f'  {yr}: {", ".join(parts) if parts else "(no production)"}  '
              f'—  {n_open} line(s) open  util={util:.1f}%')

    print(f'\nCost breakdown:')
    cost_running    = costs['running']    * n_line_years
    cost_changeover = costs['changeover'] * n_mixed_years
    total_cost      = cost_running + cost_changeover
    print(f'  Running cost  : {n_line_years} line-years '
          f'× ${costs["running"]:,.0f} = ${cost_running:,.0f}')
    print(f'  Changeov. pen.: {n_mixed_years} mixed-line-years '
          f'× ${costs["changeover"]:,.0f} = ${cost_changeover:,.0f}')
    print(f'  ─────────────────────────────────────────')
    print(f'  TOTAL         :                   ${total_cost:,.0f}')


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description='Simple ILP solver for production line allocation')
    parser.add_argument('workbook', nargs='?', default='Simple.xlsx',
                        help='Path to Simple.xlsx (default: Simple.xlsx)')
    args = parser.parse_args()

    wb_path = Path(args.workbook)
    if not wb_path.exists():
        print(f'ERROR: workbook not found: {wb_path}')
        print('Run: python create_simple_xlsx.py  to generate it first.')
        sys.exit(1)

    print(f'Loading {wb_path} ...')
    wb  = openpyxl.load_workbook(str(wb_path), data_only=True)
    inp = load_inputs(wb)

    demand_years = [yr for yr in inp['years'] if sum(inp['demand'][yr]) > 0]
    print(f'  Available seconds/year : {inp["avail"]:,.0f}')
    print(f'  Demand years           : {demand_years[0]}-{demand_years[-1]}  '
          f'({len(demand_years)} years)')
    print(f'  BASE_OEE               : {BASE_OEE*100:.0f}%')
    print(f'  CHANGEOVER_OEE_PENALTY : {CHANGEOVER_OEE_PENALTY*100:.0f}%')
    print(f'  Cost — running: ${inp["costs"]["running"]:,.0f}  '
          f'changeover: ${inp["costs"]["changeover"]:,.0f}')

    print('\nBuilding & solving ILP ...')
    result = build_and_solve(inp)
    if result is None:
        print('Solver failed — no solution found.')
        sys.exit(1)

    alloc, intro, n_line_years, n_mixed_years = result

    print('\nVerifying solution ...')
    ok = verify(alloc, inp)
    print(f'  {"All checks passed" if ok else "Issues found — see above"}')

    print_summary(alloc, intro, inp, n_line_years, n_mixed_years)

    # Reload workbook in write mode and save results
    print('\nWriting outputs to workbook ...')
    wb_write = openpyxl.load_workbook(str(wb_path))
    n_rows = write_output(wb_write, alloc, intro, inp,
                          n_line_years, n_mixed_years, ok)
    print(f'  Allocation sheet: {n_rows} rows written')

    print(f'\nSaving {wb_path} ...')
    wb_write.save(str(wb_path))
    print('Done!')


if __name__ == '__main__':
    main()

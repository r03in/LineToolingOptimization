#!/usr/bin/env python3
"""
LineToolingOptimization — ILP Solver
Replaces the greedy heuristic with a Mixed-Integer Linear Program (PuLP/CBC).

Improvements over solver_greedy.py:
  - Optimises ALL years simultaneously (global, not myopic)
  - Exploits tooling-sharing matrices (counts families, not raw products)
  - Models changeover-OEE penalty for multi-product lines (configurable)
  - Tracks named physical tooling IDs (MECH-P01 / OPTI-P01) per line per year
  - Outputs summary CSV + Gantt PNG in addition to the Excel grid

Priority order in objective:
  1. Minimise total lines open across all years
  2. Minimise tooling sets + product-line switches

Usage: python solver_ilp.py [workbook.xlsx]
Requirements: pip install pulp matplotlib openpyxl
"""
import sys
import csv
import argparse
from pathlib import Path
from collections import defaultdict

import openpyxl
import pulp
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

# ── Sheet names (Solver.xlsx layout — created by create_solver_xlsx.py) ───────
DEMAND_SHEET  = 'Demand'
PARAMS_SHEET  = 'Parameters'
TOOLING_SHEET = 'Tooling'
ALLOC_SHEET   = 'Allocation'
REPORT_SHEET  = 'Report'

NUM_PRODUCTS = 10
NUM_LINES    = 15

# 'Demand' sheet positions
DEM_DATA_ROW = 5    # first data row (year 2025); col 1=year, col 2=P1 … col 11=P10

# 'Parameters' sheet positions
PAR_HOURS_ROW    = 4    # hours/shift value in col B
PAR_SHIFTS_ROW   = 5
PAR_DAYS_ROW     = 6
PAR_WEEKS_ROW    = 7
PAR_VAL_COL      = 2
PAR_CT_DATA_ROW  = 14   # first cycle-time data row (Line 1); col 2=P1 … col 11=P10
PAR_OEE_DATA_ROW = 33   # first OEE data row (Line 1)
PAR_COST_LINE_ROW = 50  # cost of a new production line
PAR_COST_UPGR_ROW = 51  # cost of upgrading an established line with a new product
PAR_COST_MECH_ROW = 52  # cost per mechanical tooling set
PAR_COST_OPT_ROW  = 53  # cost per optical tooling set
PAR_COST_VALD_ROW = 54  # validation cost per late product introduction
PAR_COST_RUN_ROW  = 55  # annual running cost per open line (discourages premature opening)

# 'Tooling' sheet positions
TOOLING_MECH_ROW = 7    # first mech matrix data row (P1); col 2=P1 … col 11=P10
TOOLING_OPT_ROW  = 21   # first optical matrix data row (P1)

# 'Allocation' sheet output positions
ALLOC_HDR_ROW  = 5
ALLOC_DATA_ROW = 6

# ── Solver configuration (edit these to change behaviour) ─────────────────────
BASE_OEE               = 0.85   # OEE for single-product lines (default)
CHANGEOVER_OEE_PENALTY = 0.03   # OEE reduction for lines running 2+ products
SOLVER_TIME_LIMIT      = 600    # CBC wall-clock limit in seconds

# Minimum raw utilisation (units×ct / avail) a line 2+ must achieve in the year
# it first opens. Prevents commissioning a new line before existing lines are at
# capacity. Line 1 is exempt (opens for validation at tiny initial demand).
# 30 % is chosen to:
#   - Block premature openings below this threshold (e.g. a line opening at 28 %)
#   - Stay well below the ~47 % worst-case overflow of any necessary new line,
#     so the constraint never causes infeasibility when demand genuinely forces
#     a new line to open.
MIN_OPENING_UTIL = 0.30  # 30 % of raw available seconds

# Default costs (USD) — overridden by values in the Parameters sheet
DEFAULT_COST_LINE       =         0   # one-time capital: amortised into annual running cost
DEFAULT_COST_UPGRADE    =   500_000   # hardware upgrade per product added to established line
DEFAULT_COST_MECH       =   110_000   # per mechanical tooling set purchased
DEFAULT_COST_OPT        =   220_000   # per optical tooling set purchased
DEFAULT_COST_VALIDATION =   100_000   # process validation per late product intro on a line
DEFAULT_COST_RUNNING    =   500_000   # annual cost per open line ($3.5M capital / 7-yr life)

# ── Product colours for Gantt chart ───────────────────────────────────────────
PRODUCT_COLORS = [
    '#4C72B0', '#DD8452', '#55A868', '#C44E52',
    '#8172B3', '#937860', '#DA8BC3', '#8C8C8C',
    '#CCB974', '#64B5CD',
]


# ─────────────────────────────────────────────────────────────────────────────
# INPUT LOADING  (reads from Solver.xlsx multi-sheet layout)
# ─────────────────────────────────────────────────────────────────────────────
def load_inputs(wb):
    """Read all inputs from the Solver.xlsx sheet layout and return a dict."""
    ws_d = wb[DEMAND_SHEET]
    ws_p = wb[PARAMS_SHEET]
    ws_t = wb[TOOLING_SHEET]

    # Production setup → available seconds per year
    setup = {
        'hours':  float(ws_p.cell(row=PAR_HOURS_ROW,  column=PAR_VAL_COL).value or 0),
        'shifts': float(ws_p.cell(row=PAR_SHIFTS_ROW, column=PAR_VAL_COL).value or 0),
        'days':   float(ws_p.cell(row=PAR_DAYS_ROW,   column=PAR_VAL_COL).value or 0),
        'weeks':  float(ws_p.cell(row=PAR_WEEKS_ROW,  column=PAR_VAL_COL).value or 0),
    }
    avail = setup['hours'] * setup['shifts'] * setup['days'] * setup['weeks'] * 3600

    # Demand: scan rows from DEM_DATA_ROW until year column is empty
    years, demand = [], {}
    for r in range(DEM_DATA_ROW, DEM_DATA_ROW + 50):
        yr = ws_d.cell(row=r, column=1).value
        if yr is None:
            break
        yr = int(yr)
        years.append(yr)
        demand[yr] = [int(ws_d.cell(row=r, column=2 + p).value or 0)
                      for p in range(10)]

    # Cycle times and OEE: col 2 = P1 … col 11 = P10
    ct  = [[ws_p.cell(row=PAR_CT_DATA_ROW  + l, column=2 + p).value or 12
            for p in range(10)] for l in range(15)]
    oee = [[ws_p.cell(row=PAR_OEE_DATA_ROW + l, column=2 + p).value or 0.85
            for p in range(10)] for l in range(15)]

    # Tooling compatibility matrices: col 2 = P1 … col 11 = P10
    mt  = [[ws_t.cell(row=TOOLING_MECH_ROW + i, column=2 + j).value or 0
            for j in range(10)] for i in range(10)]
    ot  = [[ws_t.cell(row=TOOLING_OPT_ROW  + i, column=2 + j).value or 0
            for j in range(10)] for i in range(10)]

    # Cost parameters (fall back to defaults if cells are empty)
    def _cost(row, default):
        v = ws_p.cell(row=row, column=PAR_VAL_COL).value
        return float(v) if v is not None else default

    costs = {
        'line':       _cost(PAR_COST_LINE_ROW, DEFAULT_COST_LINE),
        'upgrade':    _cost(PAR_COST_UPGR_ROW, DEFAULT_COST_UPGRADE),
        'mech':       _cost(PAR_COST_MECH_ROW, DEFAULT_COST_MECH),
        'opt':        _cost(PAR_COST_OPT_ROW,  DEFAULT_COST_OPT),
        'validation': _cost(PAR_COST_VALD_ROW, DEFAULT_COST_VALIDATION),
        'running':    _cost(PAR_COST_RUN_ROW,  DEFAULT_COST_RUNNING),
    }

    return {'years': years, 'demand': demand, 'avail': avail,
            'days': setup['days'],
            'ct': ct, 'oee': oee, 'mt': mt, 'ot': ot, 'costs': costs}


# ─────────────────────────────────────────────────────────────────────────────
# TOOLING FAMILIES  (union-find on compatibility matrix)
# ─────────────────────────────────────────────────────────────────────────────
def compute_families(matrix, n):
    """
    Return a list of frozensets, each being a connected component of the
    product-compatibility graph defined by *matrix* (1 = compatible / shared).
    Products that share tooling belong to the same family; one physical set
    serves the whole family on a given line.
    """
    parent = list(range(n))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        parent[find(a)] = find(b)

    for i in range(n):
        for j in range(n):
            if matrix[i][j]:
                union(i, j)

    groups: dict[int, set] = defaultdict(set)
    for i in range(n):
        groups[find(i)].add(i)
    return [frozenset(s) for s in groups.values()]


# ─────────────────────────────────────────────────────────────────────────────
# ILP BUILD & SOLVE
# ─────────────────────────────────────────────────────────────────────────────
def build_and_solve(inp):
    """
    Build the Mixed-Integer Linear Program and solve with CBC.

    Returns (alloc, tooled, intro, mech_sets, opt_sets, mech_fams, opt_fams)
    or None on failure.

    alloc  : dict (year, line, product) -> int units
    tooled : list[set]  line -> set of product indices ever run on that line
    intro  : dict       line -> first year the line was used
    """
    demand = inp['demand']
    avail  = inp['avail']
    ct     = inp['ct']

    # Only model years that have actual demand — years with zero total demand
    # cause infeasibility: the no_close constraint (o[l,i] >= o[l,i-1]) would
    # force lines to stay open while constraint 4 forces o=0 when demand=0.
    years = [yr for yr in inp['years'] if sum(demand[yr]) > 0]

    P = NUM_PRODUCTS
    L = NUM_LINES
    Y = len(years)

    # Active products per year (demand > 0)
    active = {yr: [p for p in range(P) if demand[yr][p] > 0] for yr in years}

    # Big-M: max units any single product could demand in any year
    M_big = max(demand[yr][p] for yr in years for p in range(P)) + 1

    # Tooling families from compatibility matrices
    mech_fams = compute_families(inp['mt'], P)
    opt_fams  = compute_families(inp['ot'], P)
    MF = len(mech_fams)
    OF = len(opt_fams)

    # Product -> family index lookup
    prod_to_mf = {}
    for f, fam in enumerate(mech_fams):
        for p in fam:
            prod_to_mf[p] = f
    prod_to_of = {}
    for f, fam in enumerate(opt_fams):
        for p in fam:
            prod_to_of[p] = f

    prob = pulp.LpProblem('LineTooling', pulp.LpMinimize)

    # ── Decision variables ────────────────────────────────────────────────────

    # x[p,l,i]  continuous: units of product p on line l in year index i
    x = pulp.LpVariable.dicts(
        'x',
        [(p, l, i) for p in range(P) for l in range(L) for i in range(Y)],
        lowBound=0, cat='Continuous')

    # u[p,l,i]  binary: 1 if product p produced on line l in year i
    u = pulp.LpVariable.dicts(
        'u',
        [(p, l, i) for p in range(P) for l in range(L) for i in range(Y)],
        cat='Binary')

    # o[l,i]  binary: 1 if line l is open (any product) in year i
    o = pulp.LpVariable.dicts(
        'open',
        [(l, i) for l in range(L) for i in range(Y)],
        cat='Binary')

    # mul[l,i]  binary: 1 if line l runs 2+ products in year i  (OEE penalty)
    mul = pulp.LpVariable.dicts(
        'multi',
        [(l, i) for l in range(L) for i in range(Y)],
        cat='Binary')

    # tm[f,l]  binary: 1 if mech family f is ever on line l (permanent once added)
    tm = pulp.LpVariable.dicts(
        'tm',
        [(f, l) for f in range(MF) for l in range(L)],
        cat='Binary')

    # to_[f,l]  binary: 1 if optical family f is ever on line l
    to_ = pulp.LpVariable.dicts(
        'to',
        [(f, l) for f in range(OF) for l in range(L)],
        cat='Binary')

    # ever_open[l]  binary: 1 if line l is opened at any point (one-time capital)
    ever_open = pulp.LpVariable.dicts(
        'ever_open',
        [l for l in range(L)],
        cat='Binary')

    # late_intro[p,l]  binary: 1 if product p is first introduced to line l
    # after the line was already commissioned (triggers upgrade + validation cost).
    # Does NOT re-fire if the product stops and restarts — once introduced, always
    # introduced. Modelled via ever_been_u (see constraint 10).
    late_intro = pulp.LpVariable.dicts(
        'late_intro',
        [(p, l) for p in range(P) for l in range(L)],
        cat='Binary')

    # ever_been_u[p,l,i]  binary: 1 if product p has run on line l in any year
    # up to and including year index i. Non-decreasing over time.
    ever_been_u = pulp.LpVariable.dicts(
        'ever_been_u',
        [(p, l, i) for p in range(P) for l in range(L) for i in range(Y)],
        cat='Binary')

    costs = inp['costs']

    # ── Objective  (minimise total USD cost) ──────────────────────────────────
    prob += (
        costs['line']       * pulp.lpSum(ever_open[l] for l in range(L))
      + costs['running']    * pulp.lpSum(o[l, i]
                                         for l in range(L) for i in range(Y))
      + (costs['upgrade'] + costs['validation'])
                            * pulp.lpSum(late_intro[p, l]
                                         for p in range(P) for l in range(L))
      + costs['mech']      * pulp.lpSum(tm[f, l]
                                         for f in range(MF) for l in range(L))
      + costs['opt']       * pulp.lpSum(to_[f, l]
                                         for f in range(OF) for l in range(L))
    )

    # ── Constraints ───────────────────────────────────────────────────────────
    for i, yr in enumerate(years):
        d   = demand[yr]
        act = active[yr]

        # 1. Demand satisfaction: all demand must be allocated
        for p in range(P):
            prob += (pulp.lpSum(x[p, l, i] for l in range(L)) == d[p],
                     f'dem_{p}_{i}')

        # 2. If demand = 0, forbid production flags (tightens the model)
        for p in range(P):
            if d[p] == 0:
                for l in range(L):
                    prob += (u[p, l, i] == 0, f'zero_{p}_{l}_{i}')

        for l in range(L):
            # 3. Production only if flagged: x <= M * u
            for p in range(P):
                prob += (x[p, l, i] <= M_big * u[p, l, i], f'xu_{p}_{l}_{i}')

            # 4. Line open iff at least one product flag is set
            for p in range(P):
                prob += (o[l, i] >= u[p, l, i], f'open_ge_{p}_{l}_{i}')
            prob += (o[l, i] <= pulp.lpSum(u[p, l, i] for p in range(P)),
                     f'open_le_{l}_{i}')

            # 5. Multi-product flag: forced to 1 when 2+ products present
            #    sum_p u[p,l,i] - 1 <= (P-1) * mul[l,i]
            prob += (pulp.lpSum(u[p, l, i] for p in range(P)) - 1
                     <= (P - 1) * mul[l, i],
                     f'multi_{l}_{i}')
            # multi can only be 1 when line is open
            prob += (mul[l, i] <= o[l, i], f'multi_le_open_{l}_{i}')

            # 6. Capacity with linearised OEE penalty
            #    sum_p(x * ct) + avail * PENALTY * mul <= avail * BASE_OEE
            #    (mul on LHS keeps the constraint linear)
            prob += (
                pulp.lpSum(x[p, l, i] * ct[l][p] for p in range(P))
                + avail * CHANGEOVER_OEE_PENALTY * mul[l, i]
                <= avail * BASE_OEE,
                f'cap_{l}_{i}'
            )


    # 7. Line 1 validation constraint:
    #    Each product that is ever demanded must be run on Line 1 at least once
    #    across the entire planning horizon. Once run, the product is certified on
    #    that line for life — no annual re-validation required.
    #    In practice this is satisfied naturally in the first year Line 1 opens
    #    (2029, tiny demand), so it costs only tooling, no extra line capital.
    for p in range(P):
        if any(demand[yr][p] > 0 for yr in years):
            prob += (
                pulp.lpSum(u[p, 0, i] for i in range(Y)) >= 1,
                f'val1_{p}'
            )

    # 8a. Lines don't close: once a line is open it stays open.
    #     Realistic (a commissioned production line doesn't disappear) and
    #     tightens the model significantly.
    for l in range(L):
        for i in range(1, Y):
            prob += (o[l, i] >= o[l, i - 1], f'no_close_{l}_{i}')

    # 8b. Minimum opening-year utilisation for lines 2+ (Line 1 is exempt — it opens
    #     at tiny demand for initial product validation).
    #     A new line (o transitions 0→1) must reach MIN_OPENING_UTIL of raw capacity
    #     in its first year. Prevents commissioning a dedicated line just for OEE gain
    #     when existing lines still have spare capacity.
    for l in range(1, L):      # lines 2–15 only
        for i in range(1, Y):
            prob += (
                pulp.lpSum(x[p, l, i] * ct[l][p] for p in range(P))
                >= MIN_OPENING_UTIL * avail * (o[l, i] - o[l, i - 1]),
                f'min_open_util_{l}_{i}'
            )

    # 8. Tooling permanence: if product runs on line in any year, family tooling needed
    for f, fam in enumerate(mech_fams):
        for l in range(L):
            for p in fam:
                for i in range(Y):
                    prob += (tm[f, l] >= u[p, l, i], f'tm_{f}_{l}_{p}_{i}')

    for f, fam in enumerate(opt_fams):
        for l in range(L):
            for p in fam:
                for i in range(Y):
                    prob += (to_[f, l] >= u[p, l, i], f'to_{f}_{l}_{p}_{i}')

    # 9. ever_open: line counts as opened once any year-slot is active
    for l in range(L):
        for i in range(Y):
            prob += (ever_open[l] >= o[l, i], f'ever_open_{l}_{i}')

    # 10. ever_been_u: non-decreasing indicator of whether product p has ever
    #     run on line l through year i.
    #     Lower bounds: must be 1 once u=1 (and stay 1 thereafter).
    #     Upper bound:  can only be 1 if product has actually run at some point.
    for p in range(P):
        for l in range(L):
            for i in range(Y):
                # Lower: current production sets the flag
                prob += (ever_been_u[p, l, i] >= u[p, l, i],
                         f'ebu_lo_u_{p}_{l}_{i}')
                # Lower: flag is non-decreasing (once set, stays set)
                if i > 0:
                    prob += (ever_been_u[p, l, i] >= ever_been_u[p, l, i - 1],
                             f'ebu_lo_nd_{p}_{l}_{i}')
                # Upper: flag can only be 1 if product ran at some point up to i
                prob += (ever_been_u[p, l, i]
                         <= pulp.lpSum(u[p, l, j] for j in range(i + 1)),
                         f'ebu_hi_{p}_{l}_{i}')

    # 11. late_intro: fires on the FIRST introduction of product p to line l
    #     when the line was already commissioned before that introduction.
    #     Uses ever_been_u[p,l,i-1] (not u[p,l,i-1]) so that re-introduction
    #     after a production gap does NOT trigger another charge — once a product
    #     has run on a line it stays validated there for life.
    #     RHS = 1 only when: product newly appears (u=1, ever_been=0)
    #                    AND line was already open (o[l,i-1]=1).
    for p in range(P):
        for l in range(L):
            for i in range(1, Y):
                prob += (late_intro[p, l]
                         >= u[p, l, i] - ever_been_u[p, l, i - 1] + o[l, i - 1] - 1,
                         f'late_intro_{p}_{l}_{i}')

    # ── Solve ─────────────────────────────────────────────────────────────────
    n_vars  = len(prob.variables())
    n_cons  = len(prob.constraints)
    print(f'  Variables: {n_vars:,}   Constraints: {n_cons:,}')

    solver = pulp.PULP_CBC_CMD(timeLimit=SOLVER_TIME_LIMIT, msg=1)
    prob.solve(solver)

    status = pulp.LpStatus[prob.status]
    obj_val = pulp.value(prob.objective)
    print(f'  Solver status: {status}   Objective: {obj_val:,.0f}')

    # Accept Optimal (1) or time-limit with feasible (-2)
    if prob.status not in (1, -2):
        print('  ERROR: solver did not find a feasible solution.')
        return None

    # ── Extract solution ───────────────────────────────────────────────────────
    alloc  = {}
    tooled = [set() for _ in range(L)]
    intro  = {}

    for i, yr in enumerate(years):
        for l in range(L):
            for p in range(P):
                v = pulp.value(x[p, l, i])
                if v is not None and v > 0.5:
                    units = round(v)
                    alloc[(yr, l, p)] = units
                    tooled[l].add(p)
                    if l not in intro:
                        intro[l] = yr

    mech_sets = sum(
        1 for f in range(MF) for l in range(L)
        if pulp.value(tm[f, l]) is not None and pulp.value(tm[f, l]) > 0.5
    )
    opt_sets = sum(
        1 for f in range(OF) for l in range(L)
        if pulp.value(to_[f, l]) is not None and pulp.value(to_[f, l]) > 0.5
    )
    # Count lines that actually ever produced (not the ever_open variable, which
    # defaults to 1 for all lines when DEFAULT_COST_LINE=0 and costs nothing)
    n_lines_opened = len(intro)
    n_late_intros = sum(
        1 for p in range(P) for l in range(L)
        if pulp.value(late_intro[p, l]) is not None and pulp.value(late_intro[p, l]) > 0.5
    )
    n_line_years = sum(
        1 for l in range(L) for i in range(Y)
        if pulp.value(o[l, i]) is not None and pulp.value(o[l, i]) > 0.5
    )

    return (alloc, tooled, intro,
            mech_sets, opt_sets, mech_fams, opt_fams,
            n_lines_opened, n_late_intros, n_line_years)


# ─────────────────────────────────────────────────────────────────────────────
# TOOLING ID SYSTEM
# ─────────────────────────────────────────────────────────────────────────────
def compute_tooling_ids(alloc, tooled, intro, mech_fams, opt_fams, years):
    """
    For each physical tooling set (one per family per line), return a record:
      id         e.g. 'MECH-P01'  (named by lowest product number in family)
      type       'MECH' or 'OPTI'
      line       1-based line number
      intro      first year line was opened
      year_range '2031-2038'
      years      list of years the tooling was active
    """
    records = []

    for l in range(NUM_LINES):
        if not tooled[l]:
            continue

        # Mech families present on this line
        seen_mf: set[int] = set()
        for p in tooled[l]:
            for f, fam in enumerate(mech_fams):
                if p in fam:
                    seen_mf.add(f)

        for f in sorted(seen_mf):
            fam = mech_fams[f]
            rep = min(fam) + 1              # lowest product index → P-number
            tid = f'MECH-P{rep:02d}'
            active_years = [yr for yr in years
                            if any(alloc.get((yr, l, p), 0) > 0 for p in fam)]
            if active_years:
                records.append({
                    'id': tid, 'type': 'MECH', 'line': l + 1,
                    'intro': intro.get(l, '?'),
                    'years': active_years,
                    'year_range': f'{active_years[0]}-{active_years[-1]}',
                })

        # Optical families present on this line
        seen_of: set[int] = set()
        for p in tooled[l]:
            for f, fam in enumerate(opt_fams):
                if p in fam:
                    seen_of.add(f)

        for f in sorted(seen_of):
            fam = opt_fams[f]
            rep = min(fam) + 1
            tid = f'OPTI-P{rep:02d}'
            active_years = [yr for yr in years
                            if any(alloc.get((yr, l, p), 0) > 0 for p in fam)]
            if active_years:
                records.append({
                    'id': tid, 'type': 'OPTI', 'line': l + 1,
                    'intro': intro.get(l, '?'),
                    'years': active_years,
                    'year_range': f'{active_years[0]}-{active_years[-1]}',
                })

    records.sort(key=lambda r: (r['id'], r['line']))
    return records


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL OUTPUT  (formatted grid written to Allocation sheet)
# ─────────────────────────────────────────────────────────────────────────────
import datetime as _dt
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell as _MergedCell

# Colour palette
_C_TITLE  = '1F3864'
_C_SECT   = '2E75B6'
_C_COLHDR = 'BDD7EE'
_C_RONLY  = 'F2F2F2'
_C_GREEN  = 'E2EFDA'   # single-product line cell
_C_AMBER  = 'FFE699'   # mixed-product line cell
_C_WHITE  = 'FFFFFF'
_C_BORDER = 'B8CCE4'

def _ofill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def _ofont(bold=False, size=10, color='1F1F1F'):
    return Font(name='Calibri', bold=bold, size=size, color=color)

def _oborder():
    t = Side(style='thin', color=_C_BORDER)
    return Border(left=t, right=t, top=t, bottom=t)

def _oal(h='left'):
    return Alignment(horizontal=h, vertical='center')


def _write_7day_check(ws, alloc, intro, inp, used_lines, years, row_start):
    """
    Writes a what-if section below the main allocation grid:
    'If we had not opened the last line, how many of the remaining lines would
    need to work 7 days/week to absorb its demand?'

    Cells coloured red = line must extend to 7-day working.
    Returns the next free row number after the section.
    """
    P       = NUM_PRODUCTS
    avail_6 = inp['avail']
    days_6  = inp['days']
    avail_7 = avail_6 / days_6 * 7
    ct      = inp['ct']
    demand  = inp['demand']

    # ── Identify the last line added (latest intro year; highest index breaks ties)
    if len(intro) < 2:
        return row_start   # only 1 line — nothing to compare against

    last_line  = max(intro, key=lambda l: (intro[l], l))
    last_intro = intro[last_line]
    n1_lines   = [l for l in used_lines if l != last_line]
    N1         = len(n1_lines)

    years_check = [yr for yr in years if yr >= last_intro]
    if not years_check:
        return row_start

    # ── Per-year analysis ──────────────────────────────────────────────────────
    year_meta   = {}
    any_unavoid = False

    for yr in years_check:
        d       = demand[yr]
        tot_dem = sum(d)

        # Seconds the last line was using this year
        shortfall_sec = sum(
            alloc.get((yr, last_line, p), 0) * ct[last_line][p]
            for p in range(P)
        )

        # Extra capacity each n-1 line gains by going 6 → 7 days
        line_extra = {}
        for l in n1_lines:
            units_l = [alloc.get((yr, l, p), 0) for p in range(P)]
            n_prod  = sum(1 for p in range(P) if units_l[p] > 0)
            oee_eff = BASE_OEE - (CHANGEOVER_OEE_PENALTY if n_prod > 1 else 0)
            line_extra[l] = (avail_7 - avail_6) * oee_eff

        total_extra = sum(line_extra.values())
        unavoidable = shortfall_sec > 0 and total_extra < shortfall_sec

        if unavoidable:
            any_unavoid = True
            lines_7day  = set(n1_lines)
        elif shortfall_sec <= 0:
            lines_7day = set()
        else:
            # Greedy: commit highest-extra-capacity lines first
            sorted_l   = sorted(n1_lines, key=lambda l: line_extra[l], reverse=True)
            lines_7day = set()
            covered    = 0.0
            for l in sorted_l:
                if covered >= shortfall_sec:
                    break
                lines_7day.add(l)
                covered += line_extra[l]

        # Capacity in this scenario (7-day for lines_7day, 6-day for others)
        # Use demand-weighted avg cycle time for each line as proxy for unit capacity
        tot_cap = 0
        for l in n1_lines:
            units_l = [alloc.get((yr, l, p), 0) for p in range(P)]
            n_prod  = sum(1 for p in range(P) if units_l[p] > 0)
            oee_eff = BASE_OEE - (CHANGEOVER_OEE_PENALTY if n_prod > 1 else 0)
            av      = avail_7 if l in lines_7day else avail_6
            tot_ct  = sum(ct[l][p] * d[p] for p in range(P))
            avg_ct  = (tot_ct / tot_dem) if tot_dem > 0 else 12
            tot_cap += av * oee_eff / avg_ct

        year_meta[yr] = {
            'lines_7day':  lines_7day,
            'unavoidable': unavoidable,
            'tot_cap':     tot_cap,
        }

    # ── Write section to sheet ─────────────────────────────────────────────────
    C_RED    = 'FFC7CE'
    C_RED_FT = '9C0006'
    bdr      = _oborder()
    row      = row_start + 2   # blank gap

    # Banner
    ws.cell(row=row, column=1).value = (
        f'7-DAY WORKING CHECK  —  Alternative to opening L{last_line + 1}'
    )
    ws.cell(row=row, column=1).font      = Font(name='Calibri', bold=True,
                                                size=12, color=_C_WHITE)
    ws.cell(row=row, column=1).fill      = _ofill(_C_TITLE)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center',
                                                      vertical='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row,
                   end_column=max(N1 + 5, 8))
    ws.row_dimensions[row].height = 24
    row += 1

    # Subtitle
    ws.cell(row=row, column=1).value = (
        f'From {last_intro} onward: could the {N1} existing line(s) have '
        f'avoided L{last_line + 1} by working 7 days/week?  '
        f'(current schedule: {int(days_6)} days/week)'
    )
    ws.cell(row=row, column=1).font = Font(name='Calibri', italic=True,
                                            size=9, color='595959')
    ws.merge_cells(start_row=row, start_column=1, end_row=row,
                   end_column=max(N1 + 5, 8))
    ws.row_dimensions[row].height = 14
    row += 1
    ws.row_dimensions[row].height = 6
    row += 1

    # Column widths for the N1-line sub-grid
    extra_start = 2 + N1
    ws.column_dimensions['A'].width = 8
    for idx in range(N1):
        ws.column_dimensions[get_column_letter(2 + idx)].width = 14
    for i, w in enumerate([12, 14, 14, 10]):
        ws.column_dimensions[get_column_letter(extra_start + i)].width = w

    # Header
    def _h(r, c, val):
        cell = ws.cell(row=r, column=c)
        cell.value     = val
        cell.font      = Font(name='Calibri', bold=True, size=10)
        cell.fill      = _ofill(_C_COLHDR)
        cell.border    = bdr
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)

    ws.row_dimensions[row].height = 28
    _h(row, 1, 'Year')
    for idx, l in enumerate(n1_lines):
        _h(row, 2 + idx, f'L{l + 1}')
    _h(row, extra_start,     'Lines\nat 7-day')
    _h(row, extra_start + 1, 'Total\nDemand')
    _h(row, extra_start + 2, 'Capacity\n(scenario)')
    _h(row, extra_start + 3, 'Util %')
    row += 1

    # Data rows
    for yr in years_check:
        meta     = year_meta[yr]
        lines7   = meta['lines_7day']
        unavoid  = meta['unavoidable']
        d        = demand[yr]
        tot_dem  = sum(d)
        tot_cap  = meta['tot_cap']
        util_pct = (tot_dem / tot_cap * 100) if tot_cap > 0 else 0

        ws.row_dimensions[row].height = 18

        # Year cell
        yc = ws.cell(row=row, column=1)
        yc.value     = yr
        yc.font      = Font(name='Calibri', bold=True, size=10)
        yc.fill      = _ofill(C_RED if unavoid else _C_RONLY)
        yc.border    = bdr
        yc.alignment = Alignment(horizontal='center', vertical='center')

        # Line cells
        for idx, l in enumerate(n1_lines):
            col  = 2 + idx
            cell = ws.cell(row=row, column=col)
            cell.border = bdr
            prods = [p for p in range(P) if alloc.get((yr, l, p), 0) > 0]
            at_7  = l in lines7

            if not prods:
                cell.fill      = _ofill(C_RED if at_7 else 'F2F2F2')
                cell.value     = '7d —' if at_7 else '—'
                cell.font      = Font(name='Calibri', size=9, italic=True,
                                      bold=at_7,
                                      color=C_RED_FT if at_7 else 'AAAAAA')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                label = ', '.join(f'P{p+1}' for p in prods)
                multi = len(prods) > 1
                bg    = C_RED if at_7 else (_C_AMBER if multi else _C_GREEN)
                color = C_RED_FT if at_7 else '1F1F1F'
                cell.fill      = _ofill(bg)
                cell.value     = ('7d  ' if at_7 else '') + label
                cell.font      = Font(name='Calibri', size=9, bold=True,
                                      color=color)
                cell.alignment = Alignment(horizontal='center', vertical='center',
                                           wrap_text=True)

        # Summary cells
        sum_bg = C_RED if unavoid else _C_RONLY
        sum_ft = C_RED_FT if unavoid else '1F1F1F'

        def _sc(col, val, fmt=None):
            c = ws.cell(row=row, column=col)
            c.value     = val
            c.font      = Font(name='Calibri', size=10, color=sum_ft)
            c.fill      = _ofill(sum_bg)
            c.border    = bdr
            c.alignment = Alignment(horizontal='right', vertical='center')
            if fmt and isinstance(val, (int, float)):
                c.number_format = fmt

        n7_label = '⚠ NOT ENOUGH' if unavoid else len(lines7)
        _sc(extra_start,     n7_label)
        _sc(extra_start + 1, tot_dem,           '#,##0')
        _sc(extra_start + 2, int(tot_cap),      '#,##0')
        _sc(extra_start + 3, round(util_pct, 1), '0.0"%"')
        row += 1

    # Legend
    row += 1
    for bg, ft, label in [
        (C_RED,    C_RED_FT, 'Line must work 7 days/week to absorb demand from removed line'),
        (_C_GREEN, '595959', 'Line at normal 6 days/week — single product'),
        (_C_AMBER, '595959', 'Line at normal 6 days/week — multi product'),
    ]:
        c = ws.cell(row=row, column=2)
        c.fill   = _ofill(bg)
        c.value  = ''
        c.border = bdr
        c2 = ws.cell(row=row, column=3)
        c2.value = label
        c2.font  = Font(name='Calibri', italic=True, size=9, color='595959')
        ws.row_dimensions[row].height = 14
        row += 1

    # Unavoidable note
    if any_unavoid:
        row += 1
        bad_yrs = ', '.join(str(yr) for yr in years_check
                            if year_meta[yr]['unavoidable'])
        note = (
            f'\u26a0  NOTE: Even with all {N1} existing line(s) at 7 days/week, '
            f'demand in {bad_yrs} exceeds their total capacity. '
            f'Opening L{last_line + 1} (or equivalent) is unavoidable.'
        )
        c = ws.cell(row=row, column=1)
        c.value     = note
        c.font      = Font(name='Calibri', bold=True, size=10, color=C_RED_FT)
        c.fill      = _ofill('FFE7E7')
        c.alignment = Alignment(horizontal='left', vertical='center',
                                wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row,
                       end_column=max(N1 + 5, 8))
        ws.row_dimensions[row].height = 30
        row += 1

    return row


def write_output(wb, alloc, tooled, intro, inp, mech_sets, opt_sets,
                 n_line_years, n_late_intros, ok):
    """Write grid-format allocation table to the Allocation sheet in wb."""
    ws     = wb[ALLOC_SHEET]
    avail  = inp['avail']
    ct     = inp['ct']
    demand = inp['demand']
    costs  = inp['costs']

    # Active demand years only
    years = [yr for yr in inp['years'] if sum(demand[yr]) > 0]

    # Which lines were ever used?
    used_lines = sorted(intro.keys())
    N = len(used_lines)

    bdr = _oborder()

    # ── Clear everything from row 1 (handles merged title/header cells) ────────
    clear_start = 1
    clear_end   = ALLOC_DATA_ROW + 300
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row <= clear_end and rng.max_row >= clear_start:
            ws.unmerge_cells(str(rng))
    for r in range(clear_start, clear_end):
        for c in range(1, max(N + 10, 20)):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, _MergedCell):
                continue
            cell.value  = None
            cell.fill   = _ofill(_C_WHITE)
            cell.border = bdr

    # ── Title row ─────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    c1 = ws.cell(row=1, column=1)
    c1.value     = 'ALLOCATION  —  solver_ilp output'
    c1.font      = Font(name='Calibri', bold=True, size=14, color=_C_WHITE)
    c1.fill      = _ofill(_C_TITLE)
    c1.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=max(N + 5, 8))

    ts         = _dt.datetime.now().strftime('%Y-%m-%d  %H:%M')
    status_txt = '✓  All checks passed' if ok else '⚠  Issues found'
    total_sets = mech_sets + opt_sets
    ws.cell(row=2, column=1).value = (
        f'Last run: {ts}   |   {status_txt}   |   '
        f'{len(used_lines)} lines used   |   '
        f'{mech_sets} mech + {opt_sets} opt = {total_sets} tooling sets'
    )
    ws.cell(row=2, column=1).font = Font(name='Calibri', italic=True,
                                         size=9, color='595959')
    ws.merge_cells(start_row=2, start_column=1, end_row=2,
                   end_column=max(N + 5, 8))
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 6
    ws.row_dimensions[4].height = 6

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 8
    for idx in range(N):
        ws.column_dimensions[get_column_letter(2 + idx)].width = 14
    extra_start = 2 + N
    for i, w in enumerate([10, 14, 14, 12]):
        ws.column_dimensions[get_column_letter(extra_start + i)].width = w

    # ── Header row ─────────────────────────────────────────────────────────────
    def _hdr(row, col, val):
        c = ws.cell(row=row, column=col)
        c.value     = val
        c.font      = Font(name='Calibri', bold=True, size=10)
        c.fill      = _ofill(_C_COLHDR)
        c.border    = bdr
        c.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[ALLOC_HDR_ROW].height = 18
    _hdr(ALLOC_HDR_ROW, 1, 'Year')
    for idx, l in enumerate(used_lines):
        _hdr(ALLOC_HDR_ROW, 2 + idx, f'L{l+1}')
    _hdr(ALLOC_HDR_ROW, extra_start,     'Open\nLines')
    _hdr(ALLOC_HDR_ROW, extra_start + 1, 'Total\nDemand')
    _hdr(ALLOC_HDR_ROW, extra_start + 2, 'Total\nCapacity')
    _hdr(ALLOC_HDR_ROW, extra_start + 3, 'Util %')

    # ── Data rows ──────────────────────────────────────────────────────────────
    row = ALLOC_DATA_ROW
    for yr in years:
        d       = demand[yr]
        tot_dem = sum(d)
        n_open  = sum(1 for l in used_lines if l in intro and intro[l] <= yr)

        # Capacity of all open lines (OEE-adjusted)
        tot_cap = 0
        for l in used_lines:
            if l not in intro or intro[l] > yr:
                continue
            units_l = [alloc.get((yr, l, p), 0) for p in range(NUM_PRODUCTS)]
            n_prod  = sum(1 for p in range(NUM_PRODUCTS) if units_l[p] > 0)
            oee_eff = BASE_OEE - (CHANGEOVER_OEE_PENALTY if n_prod > 1 else 0)
            # Use the average cycle time for this line weighted by demand share
            tot_ct  = sum(ct[l][p] * d[p] for p in range(NUM_PRODUCTS))
            tot_d   = sum(d)
            avg_ct  = (tot_ct / tot_d) if tot_d > 0 else 12
            tot_cap += avail * oee_eff / avg_ct

        util_pct = (tot_dem / tot_cap * 100) if tot_cap > 0 else 0

        ws.row_dimensions[row].height = 18

        # Year cell
        c = ws.cell(row=row, column=1)
        c.value     = yr
        c.font      = Font(name='Calibri', bold=True, size=10)
        c.fill      = _ofill(_C_RONLY)
        c.border    = bdr
        c.alignment = Alignment(horizontal='center', vertical='center')

        # Line cells
        for idx, l in enumerate(used_lines):
            col  = 2 + idx
            cell = ws.cell(row=row, column=col)
            cell.border = bdr

            if l not in intro or intro[l] > yr:
                cell.fill  = _ofill(_C_WHITE)
                cell.value = None
                continue

            prods = [p for p in range(NUM_PRODUCTS)
                     if alloc.get((yr, l, p), 0) > 0]
            if not prods:
                # Open but idle this year
                cell.fill      = _ofill('F2F2F2')
                cell.value     = '—'
                cell.font      = Font(name='Calibri', size=9, color='AAAAAA',
                                      italic=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                label = ', '.join(f'P{p+1}' for p in prods)
                multi = len(prods) > 1
                cell.fill      = _ofill(_C_AMBER if multi else _C_GREEN)
                cell.value     = label
                cell.font      = Font(name='Calibri', size=9, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center',
                                           wrap_text=True)

        # Summary cells
        def _sum_cell(col, val, fmt=None):
            c = ws.cell(row=row, column=col)
            c.value     = val
            c.font      = Font(name='Calibri', size=10)
            c.fill      = _ofill(_C_RONLY)
            c.border    = bdr
            c.alignment = Alignment(horizontal='right', vertical='center')
            if fmt:
                c.number_format = fmt

        _sum_cell(extra_start,     n_open)
        _sum_cell(extra_start + 1, tot_dem,  '#,##0')
        _sum_cell(extra_start + 2, int(tot_cap), '#,##0')
        _sum_cell(extra_start + 3, round(util_pct, 1), '0.0"%"')
        row += 1

    # ── Colour legend ──────────────────────────────────────────────────────────
    row += 1
    for color, label in [(_C_GREEN, 'Single-product line (no changeover)'),
                         (_C_AMBER, 'Multi-product line (changeover penalty applies)'),
                         ('F2F2F2', 'Line open but idle this year')]:
        c = ws.cell(row=row, column=2)
        c.fill   = _ofill(color)
        c.value  = ''
        c.border = bdr
        c2 = ws.cell(row=row, column=3)
        c2.value = label
        c2.font  = Font(name='Calibri', italic=True, size=9, color='595959')
        ws.row_dimensions[row].height = 14
        row += 1

    # ── Lines Used summary ─────────────────────────────────────────────────────
    row += 1
    ws.cell(row=row, column=1).value = 'LINES USED'
    ws.cell(row=row, column=1).font  = Font(name='Calibri', bold=True, size=10,
                                            color=_C_WHITE)
    ws.cell(row=row, column=1).fill  = _ofill(_C_SECT)
    ws.merge_cells(start_row=row, start_column=1, end_row=row,
                   end_column=max(N + 5, 6))
    ws.row_dimensions[row].height = 18
    row += 1

    for l in used_lines:
        prods_ever = sorted(tooled[l])
        n_p        = len(prods_ever)
        oee_eff    = BASE_OEE - (CHANGEOVER_OEE_PENALTY if n_p > 1 else 0)
        prod_labels = ', '.join(f'P{p+1}' for p in prods_ever)
        label = (f'L{l+1}  intro={intro[l]}  products=[{prod_labels}]  '
                 f'OEE={oee_eff*100:.0f}%  '
                 f'tooling=[{mech_sets} mech / {opt_sets} opt total]')
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font  = Font(name='Calibri', size=10)
        ws.cell(row=row, column=1).fill  = _ofill(_C_RONLY)
        ws.merge_cells(start_row=row, start_column=1, end_row=row,
                       end_column=max(N + 5, 6))
        ws.row_dimensions[row].height = 16
        row += 1

    # ── Cost Breakdown ─────────────────────────────────────────────────────────
    row += 1
    ws.cell(row=row, column=1).value = 'COST BREAKDOWN'
    ws.cell(row=row, column=1).font  = Font(name='Calibri', bold=True, size=10,
                                            color=_C_WHITE)
    ws.cell(row=row, column=1).fill  = _ofill(_C_SECT)
    ws.merge_cells(start_row=row, start_column=1, end_row=row,
                   end_column=max(N + 5, 6))
    ws.row_dimensions[row].height = 18
    row += 1

    n_lines_opened = len(used_lines)
    cost_lines   = costs['line']       * n_lines_opened
    cost_running = costs['running']    * n_line_years
    cost_late    = (costs['upgrade'] + costs['validation']) * n_late_intros
    cost_mech    = costs['mech']      * mech_sets
    cost_opt     = costs['opt']       * opt_sets
    total_cost   = cost_lines + cost_running + cost_late + cost_mech + cost_opt

    breakdown = [
        (f'New lines: {n_lines_opened} × ${costs["line"]:,.0f}',          cost_lines),
        (f'Running:   {n_line_years} line-years × ${costs["running"]:,.0f}', cost_running),
        (f'Late intros: {n_late_intros} × ${costs["upgrade"]+costs["validation"]:,.0f}', cost_late),
        (f'Mech tooling: {mech_sets} sets × ${costs["mech"]:,.0f}',        cost_mech),
        (f'Opt tooling:  {opt_sets} sets × ${costs["opt"]:,.0f}',          cost_opt),
        ('TOTAL', total_cost),
    ]
    for label, val in breakdown:
        bold = (label == 'TOTAL')
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font  = Font(name='Calibri', bold=bold, size=10)
        ws.cell(row=row, column=1).fill  = _ofill(_C_RONLY)
        ws.merge_cells(start_row=row, start_column=1, end_row=row,
                       end_column=max(N + 3, 5))
        c = ws.cell(row=row, column=max(N + 4, 6))
        c.value         = val
        c.font          = Font(name='Calibri', bold=bold, size=10)
        c.fill          = _ofill(_C_RONLY)
        c.number_format = '$#,##0'
        c.alignment     = Alignment(horizontal='right')
        ws.row_dimensions[row].height = 16
        row += 1

    # ── 7-day working check ────────────────────────────────────────────────────
    row = _write_7day_check(ws, alloc, intro, inp, used_lines, years, row)

    return row - ALLOC_DATA_ROW


# ─────────────────────────────────────────────────────────────────────────────
# SUMMARY CSV
# ─────────────────────────────────────────────────────────────────────────────
def write_summary_csv(alloc, tooling_records, inp, intro, out_path):
    years  = inp['years']
    avail  = inp['avail']
    ct     = inp['ct']
    demand = inp['demand']

    with open(out_path, 'w', newline='') as f:
        w = csv.writer(f)

        # Section 1 – allocation per (line, year)
        w.writerow(['=== ALLOCATION ==='])
        w.writerow(['Year', 'Line', 'Products', 'Utilisation_%', 'OEE_used_%']
                   + [f'P{p+1}_units' for p in range(NUM_PRODUCTS)])

        for yr in years:
            d = demand.get(yr, [0] * 10)
            if sum(d) == 0:
                continue
            active_lines = sorted({l for (y2, l, _) in alloc if y2 == yr})
            for l in active_lines:
                units  = [alloc.get((yr, l, p), 0) for p in range(NUM_PRODUCTS)]
                prods  = [f'P{p+1}' for p in range(NUM_PRODUCTS) if units[p] > 0]
                t_used = sum(units[p] * ct[l][p] for p in range(NUM_PRODUCTS))
                util   = t_used / avail * 100
                n_prod = len(prods)
                oee    = (BASE_OEE - CHANGEOVER_OEE_PENALTY if n_prod > 1
                          else BASE_OEE) * 100
                w.writerow([yr, f'L{l+1}', '+'.join(prods),
                             f'{util:.1f}', f'{oee:.0f}'] + units)

        w.writerow([])
        # Section 2 – tooling ID table
        w.writerow(['=== TOOLING IDs ==='])
        w.writerow(['Tooling_ID', 'Type', 'Line', 'Intro_Year', 'Years_Active'])
        for r in tooling_records:
            w.writerow([r['id'], r['type'], f"L{r['line']}", r['intro'],
                        r['year_range']])

    print(f'  Summary CSV: {out_path}')


# ─────────────────────────────────────────────────────────────────────────────
# GANTT CHART
# ─────────────────────────────────────────────────────────────────────────────
def plot_gantt(alloc, inp, intro, out_path):
    years = inp['years']

    active_lines = sorted({l for (_, l, _) in alloc})
    if not active_lines:
        return

    fig_h = max(4, len(active_lines) * 1.4 + 2)
    fig, ax = plt.subplots(figsize=(18, fig_h))
    ax.set_title('Production Line Allocation — Gantt', fontsize=14, pad=12)
    ax.set_xlabel('Year', fontsize=11)
    ax.set_ylabel('Line', fontsize=11)

    bar_height = 0.65
    yticks, ylabels = [], []

    for row_idx, l in enumerate(active_lines):
        yticks.append(row_idx)
        ylabels.append(f'L{l+1}')

        for yr in years:
            units = [alloc.get((yr, l, p), 0) for p in range(NUM_PRODUCTS)]
            total = sum(units)
            if total == 0:
                continue

            left = yr - 0.4
            for p in range(NUM_PRODUCTS):
                if units[p] == 0:
                    continue
                width = units[p] / total * 0.8  # 0.8 = bar width per year cell
                ax.barh(row_idx, width, left=left, height=bar_height,
                        color=PRODUCT_COLORS[p % len(PRODUCT_COLORS)], alpha=0.85,
                        edgecolor='white', linewidth=0.4)
                left += width

            # Mark intro year
            if intro.get(l) == yr:
                ax.annotate(
                    f'intro {yr}',
                    xy=(yr, row_idx + bar_height / 2 + 0.05),
                    xytext=(0, 3), textcoords='offset points',
                    fontsize=6.5, ha='center', va='bottom', color='#333333',
                )

    ax.set_yticks(yticks)
    ax.set_yticklabels(ylabels, fontsize=10)
    ax.set_xticks(years)
    ax.set_xticklabels([str(y) for y in years], rotation=45, ha='right', fontsize=9)
    ax.set_xlim(min(years) - 0.6, max(years) + 0.6)
    ax.set_ylim(-0.6, len(active_lines) - 0.4)

    legend_patches = [
        mpatches.Patch(color=PRODUCT_COLORS[p], label=f'P{p+1}', alpha=0.85)
        for p in range(NUM_PRODUCTS)
    ]
    ax.legend(handles=legend_patches, loc='upper right', ncol=5,
              fontsize=8, framealpha=0.9)
    ax.grid(axis='x', alpha=0.25, linestyle='--')
    ax.invert_yaxis()

    plt.tight_layout()
    plt.savefig(out_path, dpi=150)
    plt.close()
    print(f'  Gantt chart: {out_path}')


# ─────────────────────────────────────────────────────────────────────────────
# REPORT SHEET
# ─────────────────────────────────────────────────────────────────────────────
def write_report_sheet(wb, alloc, tooled, intro, mech_fams, opt_fams,
                       mech_sets, opt_sets, n_lines_opened, n_late_intros,
                       inp, ok):
    """Write lines summary, tooling ID registry, and demand validation."""
    import datetime
    ws     = wb[REPORT_SHEET]
    years  = inp['years']
    demand = inp['demand']
    ct     = inp['ct']
    avail  = inp['avail']

    # Clear everything from row 2 downward; unmerge first to avoid MergedCell errors.
    from openpyxl.cell import MergedCell
    for rng in list(ws.merged_cells.ranges):
        if rng.max_row >= 2:
            ws.unmerge_cells(str(rng))
    for r in range(2, 200):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            cell.value  = None
            cell.fill   = _ofill('FFFFFF')
            cell.border = _oborder()

    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    def _rfill(h): return PatternFill('solid', fgColor=h)
    def _rfont(bold=False, sz=10, color='1F1F1F'):
        return Font(name='Calibri', bold=bold, size=sz, color=color)
    def _rbd():
        t = Side(style='thin', color='B8CCE4')
        return Border(left=t, right=t, top=t, bottom=t)
    def _ral(h='left'):
        return Alignment(horizontal=h, vertical='center')
    def _sect_r(row, text, ncols=7):
        ws.cell(row=row, column=1).value = text
        ws.cell(row=row, column=1).font  = Font(name='Calibri', bold=True,
                                                size=11, color='FFFFFF')
        for c in range(1, ncols + 1):
            ws.cell(row=row, column=c).fill = _rfill('2E75B6')
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=ncols)
        ws.row_dimensions[row].height = 20
    def _hdr_r(row, labels):
        for i, lbl in enumerate(labels):
            c = ws.cell(row=row, column=1 + i)
            c.value, c.font, c.fill, c.border, c.alignment = \
                lbl, _rfont(bold=True), _rfill('BDD7EE'), _rbd(), _ral('center')
        ws.row_dimensions[row].height = 18
    def _data_r(row, vals, bg='E2EFDA'):
        for i, v in enumerate(vals):
            c = ws.cell(row=row, column=1 + i)
            c.value, c.fill, c.font, c.border = v, _rfill(bg), _rfont(), _rbd()
            c.alignment = _ral('right' if isinstance(v, (int, float)) else 'left')
        ws.row_dimensions[row].height = 15

    costs = inp['costs']
    total_cost = (
        costs['line']       * n_lines_opened
      + (costs['upgrade'] + costs['validation']) * n_late_intros
      + costs['mech']      * mech_sets
      + costs['opt']       * opt_sets
    )

    ts = datetime.datetime.now().strftime('%Y-%m-%d  %H:%M')
    ws.cell(row=2, column=1).value = (
        f'Last run: {ts}   |   '
        f'{"✓  All OK" if ok else "⚠  Issues found"}   |   '
        f'{mech_sets} mech + {opt_sets} opt = {mech_sets + opt_sets} sets   |   '
        f'Total cost: ${total_cost:,.0f}'
    )
    ws.cell(row=2, column=1).font = Font(name='Calibri', italic=True,
                                         size=10, color='595959')
    ws.merge_cells('A2:G2')
    ws.row_dimensions[3].height = 6

    # ── Cost breakdown ────────────────────────────────────────────────────────
    _sect_r(4, '  COST BREAKDOWN')
    _hdr_r(5, ['Cost item', 'Unit cost (USD)', 'Qty', 'Total (USD)'])
    cost_items = [
        ('New production lines',        costs['line'],       n_lines_opened,
         costs['line'] * n_lines_opened),
        ('Line upgrades (new products)', costs['upgrade'],   n_late_intros,
         costs['upgrade'] * n_late_intros),
        ('Validation events',            costs['validation'], n_late_intros,
         costs['validation'] * n_late_intros),
        ('Mechanical tooling sets',      costs['mech'],      mech_sets,
         costs['mech'] * mech_sets),
        ('Optical tooling sets',         costs['opt'],       opt_sets,
         costs['opt'] * opt_sets),
    ]
    for ci, (label, unit_cost, qty, subtotal) in enumerate(cost_items):
        bg = 'F0F8F0' if ci % 2 else 'E2EFDA'
        _data_r(6 + ci, [label, unit_cost, qty, subtotal], bg=bg)
    # Total row
    total_row = 6 + len(cost_items)
    _data_r(total_row,
            ['TOTAL', '', '', total_cost],
            bg='BDD7EE')
    ws.cell(row=total_row, column=1).font = _rfont(bold=True, sz=10)
    ws.cell(row=total_row, column=4).font = _rfont(bold=True, sz=10)
    for col in [2, 4]:
        ws.cell(row=total_row - 5, column=col).number_format = '$#,##0'
    for ci in range(len(cost_items)):
        for col in [2, 4]:
            ws.cell(row=6 + ci, column=col).number_format = '$#,##0'
    ws.cell(row=total_row, column=4).number_format = '$#,##0'

    r = total_row + 2  # spacer before next section

    # ── Lines summary ─────────────────────────────────────────────────────────
    _sect_r(r, '  LINES SUMMARY'); r += 1
    _hdr_r(r, ['Line', 'Intro', 'Products', 'Tooling sets', 'Eff. OEE %',
               'Peak util %', 'Peak year']); r += 1
    for l in sorted(intro.keys()):
        ps     = sorted(f'P{p+1}' for p in tooled[l])
        n_prod = len(ps)
        oee    = int(round((BASE_OEE - CHANGEOVER_OEE_PENALTY
                            if n_prod > 1 else BASE_OEE) * 100))
        # Count tooling sets for this line
        mf_cnt = sum(1 for f, fam in enumerate(mech_fams) if fam & tooled[l])
        of_cnt = sum(1 for f, fam in enumerate(opt_fams)  if fam & tooled[l])
        # Peak utilisation
        peak_util, peak_yr = 0.0, ''
        for yr in years:
            d = demand.get(yr, [0] * 10)
            if sum(d) == 0:
                continue
            t = sum(alloc.get((yr, l, p), 0) * ct[l][p] for p in range(NUM_PRODUCTS))
            u = t / avail * 100
            if u > peak_util:
                peak_util, peak_yr = u, yr
        bg = 'F0F8F0' if r % 2 else 'E2EFDA'
        _data_r(r, [f'L{l+1}', intro[l], '+'.join(ps),
                    f'{mf_cnt}m + {of_cnt}o', oee,
                    round(peak_util, 1), peak_yr], bg=bg)
        r += 1

    r += 1  # spacer
    # ── Tooling ID registry ───────────────────────────────────────────────────
    _sect_r(r, '  TOOLING ID REGISTRY'); r += 1
    _hdr_r(r, ['Tooling ID', 'Type', 'Line', 'Intro year', 'Active years']); r += 1
    tooling_records = compute_tooling_ids(
        alloc, tooled, intro, mech_fams, opt_fams, years)
    for i, rec in enumerate(tooling_records):
        bg = 'F0F8F0' if i % 2 else 'E2EFDA'
        _data_r(r, [rec['id'], rec['type'], f"L{rec['line']}",
                    rec['intro'], rec['year_range']], bg=bg)
        r += 1

    r += 1  # spacer
    # ── Demand validation ─────────────────────────────────────────────────────
    _sect_r(r, '  DEMAND VALIDATION'); r += 1
    _hdr_r(r, ['Year', 'Product', 'Demand', 'Allocated', 'Difference', 'Status']); r += 1
    issues = 0
    for yr in years:
        d = demand.get(yr, [0] * 10)
        if sum(d) == 0:
            continue
        for p in range(NUM_PRODUCTS):
            alloc_p = sum(alloc.get((yr, l, p), 0) for l in range(NUM_LINES))
            diff    = alloc_p - d[p]
            if d[p] == 0 and alloc_p == 0:
                continue
            ok_row = abs(diff) <= 1
            if not ok_row:
                issues += 1
            bg = 'E2EFDA' if ok_row else 'FFC7CE'
            _data_r(r, [yr, f'P{p+1}', d[p], alloc_p, diff,
                        'OK' if ok_row else 'MISMATCH'], bg=bg)
            r += 1
    if issues == 0:
        ws.cell(row=r, column=1).value = '✓  All demand satisfied within ±1 unit tolerance.'
        ws.cell(row=r, column=1).font = Font(name='Calibri', italic=True,
                                             size=10, color='2E75B6')
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)


# ─────────────────────────────────────────────────────────────────────────────
# VERIFICATION  (demand met & capacity not exceeded)
# ─────────────────────────────────────────────────────────────────────────────
def verify(alloc, inp):
    ok = True
    for yr in inp['years']:
        d = inp['demand'].get(yr, [0] * 10)
        if sum(d) == 0:
            continue

        # Demand check
        ta = [0] * 10
        for l in range(NUM_LINES):
            for p in range(10):
                ta[p] += alloc.get((yr, l, p), 0)
        for p in range(10):
            if abs(d[p] - ta[p]) > 1:
                print(f'  X {yr} P{p+1}: demand={d[p]:,}  alloc={ta[p]:,}')
                ok = False

        # Capacity check: multi-product lines use the reduced OEE capacity
        for l in range(NUM_LINES):
            units_l = [alloc.get((yr, l, p), 0) for p in range(10)]
            t = sum(units_l[p] * inp['ct'][l][p] for p in range(10))
            n_prod = sum(1 for p in range(10) if units_l[p] > 0)
            oee_eff = (BASE_OEE - CHANGEOVER_OEE_PENALTY if n_prod > 1 else BASE_OEE)
            cap = inp['avail'] * oee_eff
            if t > cap * 1.001:
                print(f'  X {yr} Line {l+1}: capacity exceeded '
                      f'({t:,.0f} > {cap:,.0f}, OEE={oee_eff*100:.0f}%)')
                ok = False

    return ok


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description='Production line ILP optimizer')
    parser.add_argument('workbook', nargs='?', default='Solver.xlsx')
    args = parser.parse_args()

    path = Path(args.workbook)
    if not path.exists():
        print(f'Error: {path} not found')
        print('  Run  python create_solver_xlsx.py  to generate Solver.xlsx first.')
        sys.exit(1)

    print(f'Loading {path} ...')
    wb  = openpyxl.load_workbook(str(path))
    inp = load_inputs(wb)

    years = inp['years']
    print(f'  Available seconds/year : {inp["avail"]:,.0f}')
    print(f'  Years                  : {years[0]}-{years[-1]}  ({len(years)} years)')
    print(f'  BASE_OEE               : {BASE_OEE*100:.0f}%')
    print(f'  CHANGEOVER_OEE_PENALTY : {CHANGEOVER_OEE_PENALTY*100:.0f}%')

    mf_preview = compute_families(inp['mt'], NUM_PRODUCTS)
    of_preview = compute_families(inp['ot'], NUM_PRODUCTS)
    print(f'  Mech tooling families  : {len(mf_preview)}'
          f'  (sizes: {sorted(len(f) for f in mf_preview)})')
    print(f'  Opt  tooling families  : {len(of_preview)}'
          f'  (sizes: {sorted(len(f) for f in of_preview)})')
    c = inp['costs']
    print(f'  Cost — line: ${c["line"]:,.0f}  upgrade: ${c["upgrade"]:,.0f}  '
          f'mech: ${c["mech"]:,.0f}  opt: ${c["opt"]:,.0f}  '
          f'validation: ${c["validation"]:,.0f}')

    print('\nBuilding & solving ILP ...')
    result = build_and_solve(inp)
    if result is None:
        print('Solver failed — no solution found.')
        sys.exit(1)

    alloc, tooled, intro, mech_sets, opt_sets, mech_fams, opt_fams, \
        n_lines_opened, n_late_intros, n_line_years = result

    print('\nVerifying solution ...')
    ok = verify(alloc, inp)
    print(f'  {"All checks passed" if ok else "Issues found — see above"}')

    total_sets = mech_sets + opt_sets
    print(f'\nTooling: {mech_sets} mech + {opt_sets} opt = {total_sets} sets')
    print(f'  (greedy baseline: 19+19=38  |  theory min: 14+14=28)')

    costs = inp['costs']
    cost_lines   = costs['line']       * n_lines_opened
    cost_running = costs['running']    * n_line_years
    cost_late    = (costs['upgrade'] + costs['validation']) * n_late_intros
    cost_mech    = costs['mech']      * mech_sets
    cost_opt     = costs['opt']       * opt_sets
    total_cost   = cost_lines + cost_running + cost_late + cost_mech + cost_opt
    print(f'\nCost breakdown:')
    print(f'  Lines opened      : {n_lines_opened}  × ${costs["line"]:,.0f}  = ${cost_lines:,.0f}')
    print(f'  Running cost      : {n_line_years} line-years × ${costs["running"]:,.0f}  = ${cost_running:,.0f}')
    print(f'  Late intros       : {n_late_intros}  × ${costs["upgrade"]+costs["validation"]:,.0f}  = ${cost_late:,.0f}')
    print(f'  Mech tooling sets : {mech_sets}  × ${costs["mech"]:,.0f}  = ${cost_mech:,.0f}')
    print(f'  Opt  tooling sets : {opt_sets}  × ${costs["opt"]:,.0f}  = ${cost_opt:,.0f}')
    print(f'  ─────────────────────────────────────────────────')
    print(f'  TOTAL             :                    ${total_cost:,.0f}')

    print('\nLines used:')
    for l in sorted(intro.keys()):
        ps    = sorted(f'P{p+1}' for p in tooled[l])
        n_p   = len(ps)
        oee_e = (BASE_OEE - CHANGEOVER_OEE_PENALTY if n_p > 1 else BASE_OEE) * 100
        print(f'  Line {l+1:2d}  intro={intro[l]}  products={ps}'
              f'  effective_OEE={oee_e:.0f}%')

    print('\nPer-year summary:')
    for yr in years:
        d = inp['demand'].get(yr, [0] * 10)
        if sum(d) == 0:
            continue
        active_lines = sorted({l for (y2, l, _) in alloc if y2 == yr})
        parts = [f'L{l+1}=[{",".join(f"P{p+1}" for p in range(NUM_PRODUCTS) if alloc.get((yr,l,p),0)>0)}]'
                 for l in active_lines]
        print(f'  {yr}: {", ".join(parts)}  —  {len(active_lines)} line(s)')

    print('\nWriting outputs to workbook ...')
    outdir = path.parent

    n_rows = write_output(wb, alloc, tooled, intro, inp, mech_sets, opt_sets,
                          n_line_years, n_late_intros, ok)
    print(f'  Allocation sheet: {n_rows} rows written')

    write_report_sheet(wb, alloc, tooled, intro, mech_fams, opt_fams,
                       mech_sets, opt_sets, n_lines_opened, n_late_intros,
                       inp, ok)
    print('  Report sheet: written')

    tooling_records = compute_tooling_ids(
        alloc, tooled, intro, mech_fams, opt_fams, years)
    write_summary_csv(alloc, tooling_records, inp, intro,
                      str(outdir / 'tooling_summary.csv'))
    plot_gantt(alloc, inp, intro, str(outdir / 'line_gantt.png'))

    print(f'\nSaving {path} ...')
    wb.save(str(path))
    print('Done!')


if __name__ == '__main__':
    main()
